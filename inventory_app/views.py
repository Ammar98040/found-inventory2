from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db import models as db_models
from .models import Product, Location, Warehouse, AuditLog, Order, ProductReturn, UserProfile, UserActivityLog, Container, SecureBackup
from .decorators import admin_required, staff_required, exclude_maintenance, exclude_admin_dashboard, get_user_type, is_admin
from .forms import LoginForm, RegisterStaffForm, ProductForm, EditStaffForm
from .ocr_service import analyze_invoice_image
import json
import logging
from django.core import serializers
from datetime import datetime, timedelta
from django.utils import timezone
from django.views.decorators.cache import never_cache, cache_page
from django.core.cache import cache
from django.core.paginator import Paginator
from django.conf import settings
from django.utils.deprecation import MiddlewareMixin

# إعداد Logger للأمان
security_logger = logging.getLogger('inventory_app.security')
logger = logging.getLogger('inventory_app')

# Helper function للـ Caching
def get_cached_or_set(cache_key, callable_func, timeout=300):
    """
    الحصول على البيانات من Cache أو تنفيذ الدالة وحفظها في Cache
    
    Args:
        cache_key: مفتاح Cache
        callable_func: الدالة التي تُنفذ إذا لم تكن البيانات في Cache
        timeout: مدة الحفظ في Cache بالثواني (افتراضي: 5 دقائق)
    
    Returns:
        البيانات من Cache أو من تنفيذ الدالة
    """
    data = cache.get(cache_key)
    if data is None:
        data = callable_func()
        cache.set(cache_key, data, timeout)
    return data

# Rate limiting - استخدام django-ratelimit إذا كان متاحاً
try:
    from django_ratelimit.decorators import ratelimit  # pyright: ignore[reportMissingImports]
    RATELIMIT_AVAILABLE = True
except ImportError:
    RATELIMIT_AVAILABLE = False
    # إنشاء decorator بديل إذا لم يكن django-ratelimit متاحاً
    def ratelimit(*args, **kwargs):
        def decorator(func):
            return func
        return decorator


@login_required
def home(request):
    """الصفحة الرئيسية - البحث عن المنتجات"""
    user = request.user
    user_profile = None
    user_type_display = 'موظف'
    
    if hasattr(user, 'user_profile'):
        user_profile = user.user_profile
        user_type_display = user_profile.get_user_type_display()
    elif user.is_superuser:
        user_type_display = 'مسؤول'
    
    return render(request, 'inventory_app/home.html', {
        'user': user,
        'user_profile': user_profile,
        'user_type_display': user_type_display
    })


@csrf_exempt
@transaction.atomic
def confirm_products(request):
    """تأكيد أخذ المنتجات وخصم الكميات - آمن وذري ومتحقق بالكامل"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'بيانات غير صالحة (JSON)'}, status=400)

    products_list = data.get('products', [])
    recipient_name = str(data.get('recipient_name', '') or '').strip()

    if not isinstance(products_list, list) or len(products_list) == 0:
        return JsonResponse({'success': False, 'error': 'لم يتم تحديد أي منتجات للسحب'}, status=400)

    # تجميع الطلبات لنفس المنتج لتفادي التكرارات
    aggregated_requests = {}
    invalid_items = []
    
    # التحقق من أن جميع المنتجات لها كمية أكبر من 0
    zero_quantity_items = []
    
    for item in products_list:
        try:
            number = str(item.get('number', '') or '').strip()
            # التأكد من تحويل الكمية إلى رقم صحيح، وإذا كانت فارغة تعتبر 0
            qty_raw = item.get('quantity')
            if qty_raw is None or qty_raw == '':
                qty = 0
            else:
                qty = int(qty_raw)
        except (ValueError, TypeError):
            invalid_items.append(item)
            continue
            
        if not number:
            invalid_items.append(item)
            continue
            
        if qty <= 0:
            zero_quantity_items.append(number)
            continue
            
        aggregated_requests[number] = aggregated_requests.get(number, 0) + qty

    if invalid_items:
        return JsonResponse({'success': False, 'error': 'مدخلات غير صالحة', 'invalid_items': invalid_items}, status=400)
        
    if zero_quantity_items:
        items_str = ", ".join(zero_quantity_items)
        example_str = f"{zero_quantity_items[0]}:5" if zero_quantity_items else "رقم_المنتج:الكمية"
        return JsonResponse({
            'success': False, 
            'error': f'يجب تحديد الكمية للمنتجات التالية: {items_str}. \nالرجاء البحث باستخدام الصيغة: {example_str}', 
            'zero_quantity_items': zero_quantity_items
        }, status=400)

    if not aggregated_requests:
         return JsonResponse({'success': False, 'error': 'لم يتم تحديد أي كميات للسحب'}, status=400)

    product_numbers = sorted(aggregated_requests.keys())

    # قفل صفوف المنتجات بترتيب ثابت لتفادي الـ deadlocks
    products = list(
        Product.objects.filter(product_number__in=product_numbers)
        .select_for_update()
        .order_by('product_number')
    )
    products_dict = {p.product_number: p for p in products}

    # التحقق من وجود جميع المنتجات
    missing = [n for n in product_numbers if n not in products_dict]
    if missing:
        return JsonResponse({'success': False, 'error': 'منتجات غير موجودة', 'missing': missing}, status=404)

    # التحقق من الكميات المتاحة لجميع المنتجات قبل أي خصم
    insufficient = []
    for n in product_numbers:
        requested = aggregated_requests[n]
        product = products_dict[n]
        if requested == 0:
            continue
        if product.quantity < requested:
            insufficient.append({'product_number': n, 'available': product.quantity, 'requested': requested})

    if insufficient:
        return JsonResponse({'success': False, 'error': 'كميات غير كافية', 'insufficient': insufficient}, status=400)

    # الخصم الذري مع تسجيل دقيق
    updated_products = []
    for n in product_numbers:
        requested = aggregated_requests[n]
        product = products_dict[n]
        old_quantity = product.quantity

        if requested == 0:
            AuditLog.objects.create(
                action='quantity_taken',
                product=product,
                product_number=n,
                quantity_before=old_quantity,
                quantity_after=old_quantity,
                quantity_change=0,
                notes='تم تأكيد بدون سحب (كمية 0)',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            updated_products.append({
                'product_number': n,
                'old_quantity': old_quantity,
                'new_quantity': old_quantity,
                'quantity_taken': 0
            })
            continue

        # خصم آمن
        product.quantity = old_quantity - requested
        product.save(update_fields=['quantity'])

        AuditLog.objects.create(
            action='quantity_taken',
            product=product,
            product_number=n,
            quantity_before=old_quantity,
            quantity_after=product.quantity,
            quantity_change=-requested,
            notes=f'خصم دفعة: {requested}',
            user=request.user.username if request.user.is_authenticated else 'Guest'
        )

        updated_products.append({
            'product_number': n,
            'name': product.name,
            'category': product.category,
            'old_quantity': old_quantity,
            'new_quantity': product.quantity,
            'quantity_taken': requested
        })

    # إنشاء الطلبية بشكل ذري بعد نجاح الخصم للجميع
    from datetime import datetime
    import random
    import string
    order_number = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}-{''.join(random.choices(string.ascii_uppercase + string.digits, k=4))}"

    total_products = len([p for p in updated_products if p['quantity_taken'] > 0])
    total_quantities = sum(p['quantity_taken'] for p in updated_products)

    Order.objects.create(
        order_number=order_number,
        products_data=updated_products,
        total_products=total_products,
        total_quantities=total_quantities,
        recipient_name=recipient_name or None,
        user=request.user.username if request.user.is_authenticated else 'Guest'
    )

    return JsonResponse({
        'success': True,
        'updated_products': updated_products,
        'message': f'تم خصم {total_products} منتج بإجمالي {total_quantities}',
        'order_number': order_number
    })


@csrf_exempt
def search_products(request):
    """البحث عن المنتجات من خلال أرقامهم"""
    if request.method == 'POST':
        data = json.loads(request.body)
        products_list = data.get('products', [])
        semantic = data.get('semantic', True)
        
        # Extract product numbers for efficient batch query
        product_numbers = [item.get('product_number', '').strip() for item in products_list if item.get('product_number', '').strip()]
        
        # Optimize: Use select_related to avoid N+1 queries and batch query
        products = Product.objects.filter(product_number__in=product_numbers).select_related('location')
        
        # Create a dictionary for fast lookup
        products_dict = {p.product_number: p for p in products}
        
        results = []
        
        for item in products_list:
            product_number = item.get('product_number', '').strip()
            requested_quantity = int(item.get('quantity', 0))
            
            if not product_number:
                continue
            
            # Fast lookup from dictionary
            product = products_dict.get(product_number)
            
            if product:
                locations_data = []
                if product.location:
                    grid_pos = product.location.get_grid_position()
                    locations_data.append({
                        'id': product.location.id,
                        'full_location': product.location.full_location,
                        'row': product.location.row,
                        'column': product.location.column,
                        'x': grid_pos['x'],
                        'y': grid_pos['y'],
                        'notes': product.location.notes,
                    })
                
                result = {
                    'product_number': product.product_number,
                    'name': product.name,
                    'category': product.category,
                    'quantity': product.quantity,
                    'locations': locations_data,
                    'found': True,
                }
                
                if requested_quantity > 0:
                    result['requested_quantity'] = requested_quantity
                
                results.append(result)
            else:
                suggestions = []
                if semantic:
                    from difflib import SequenceMatcher
                    from django.db.models import Q
                    candidates_qs = Product.objects.filter(
                        Q(product_number__icontains=product_number) | Q(name__icontains=product_number)
                    )[:20]
                    scored = []
                    for p in candidates_qs:
                        s1 = SequenceMatcher(None, product_number, p.product_number).ratio()
                        s2 = SequenceMatcher(None, product_number, p.name or '').ratio()
                        score = max(s1, s2)
                        scored.append((score, p))
                    scored.sort(key=lambda x: x[0], reverse=True)
                    for score, p in scored[:5]:
                        suggestions.append({
                            'product_number': p.product_number,
                            'name': p.name,
                            'category': p.category,
                            'quantity': p.quantity
                        })
                results.append({
                    'product_number': product_number,
                    'requested_quantity': requested_quantity,
                    'found': False,
                    'error': 'المنتج غير موجود في قاعدة البيانات',
                    'suggestions': suggestions
                })
        
        return JsonResponse({'results': results}, json_dumps_params={'ensure_ascii': False})
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@require_http_methods(["GET"])
def get_products_list(request):
    products = Product.objects.all()[:100]
    products_data = [{'number': p.product_number, 'name': p.name} for p in products]
    return JsonResponse({'products': products_data}, json_dumps_params={'ensure_ascii': False})


def manage_warehouse(request):
    """صفحة إدارة المستودع"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        warehouse = Warehouse.objects.create(name='المستودع الرئيسي', rows_count=6, columns_count=15)
    
    return render(request, 'inventory_app/manage_warehouse.html', {'warehouse': warehouse})


@require_http_methods(["POST"])
@login_required
def compact_row(request):
    """ضغط الصف (إزالة الفراغات)"""
    try:
        data = json.loads(request.body)
        row_num = int(data.get('row'))
        
        warehouse = Warehouse.objects.first()
        if not warehouse:
            return JsonResponse({'success': False, 'error': 'المستودع غير موجود'})
            
        with transaction.atomic():
            # الحصول على جميع مواقع الصف مرتبة حسب العمود
            locations = list(Location.objects.filter(
                warehouse=warehouse, 
                row=row_num
            ).order_by('column').prefetch_related('products'))
            
            # حفظ الحالة السابقة للتراجع
            undo_data = []
            for loc in locations:
                for product in loc.products.all():
                    undo_data.append({
                        'product_id': product.id,
                        'location_id': loc.id
                    })
            
            # حفظ في الجلسة
            request.session['last_compaction_undo'] = {
                'type': 'row',
                'id': row_num,
                'data': undo_data,
                'timestamp': timezone.now().isoformat()
            }
            
            # تجميع المنتجات من المواقع المشغولة
            occupied_groups = []
            for loc in locations:
                products = list(loc.products.all())
                if products:
                    occupied_groups.append(products)
            
            # إعادة توزيع المنتجات على المواقع الأولى
            for i, group in enumerate(occupied_groups):
                target_location = locations[i]
                
                for product in group:
                    if product.location_id != target_location.id:
                        product.location = target_location
                        product.save(update_fields=['location'])
                        
        return JsonResponse({'success': True, 'message': f'تم إعادة ترتيب الصف {row_num} بنجاح', 'can_undo': True})
        
    except ValueError:
        return JsonResponse({'success': False, 'error': 'بيانات غير صالحة'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})


@csrf_exempt
@require_http_methods(["POST"])
def analyze_image_view(request):
    try:
        if 'image' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No image provided'}, status=400)
            
        image_file = request.FILES['image']
        
        # Call the OCR service
        result = analyze_invoice_image(image_file)
        
        if isinstance(result, list):
            return JsonResponse({'success': True, 'data': {'products': result}})
        elif isinstance(result, dict) and 'error' in result:
            return JsonResponse({'success': False, 'error': result['error']}, status=500)
        else:
            return JsonResponse({'success': False, 'error': 'Failed to analyze image'}, status=500)
            
    except Exception as e:
        logger.error(f"Error in analyze_image_view: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
@login_required
def compact_column(request):
    """ضغط العمود (إزالة الفراغات)"""
    try:
        data = json.loads(request.body)
        col_num = int(data.get('column'))
        
        warehouse = Warehouse.objects.first()
        if not warehouse:
            return JsonResponse({'success': False, 'error': 'المستودع غير موجود'})
            
        with transaction.atomic():
            # الحصول على جميع مواقع العمود مرتبة حسب الصف
            locations = list(Location.objects.filter(
                warehouse=warehouse, 
                column=col_num
            ).order_by('row').prefetch_related('products'))
            
            # حفظ الحالة السابقة للتراجع
            undo_data = []
            for loc in locations:
                for product in loc.products.all():
                    undo_data.append({
                        'product_id': product.id,
                        'location_id': loc.id
                    })
            
            # حفظ في الجلسة
            request.session['last_compaction_undo'] = {
                'type': 'column',
                'id': col_num,
                'data': undo_data,
                'timestamp': timezone.now().isoformat()
            }
            
            # تجميع المنتجات من المواقع المشغولة
            occupied_groups = []
            for loc in locations:
                products = list(loc.products.all())
                if products:
                    occupied_groups.append(products)
            
            # إعادة توزيع المنتجات على المواقع الأولى (من الأعلى للأسفل)
            for i, group in enumerate(occupied_groups):
                target_location = locations[i]
                
                for product in group:
                    if product.location_id != target_location.id:
                        product.location = target_location
                        product.save(update_fields=['location'])
                        
        return JsonResponse({'success': True, 'message': f'تم إعادة ترتيب العمود {col_num} بنجاح', 'can_undo': True})
        
    except ValueError:
        return JsonResponse({'success': False, 'error': 'بيانات غير صالحة'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})


@require_http_methods(["POST"])
@login_required
def revert_compaction(request):
    """التراجع عن آخر عملية ترتيب"""
    try:
        undo_info = request.session.get('last_compaction_undo')
        if not undo_info:
            return JsonResponse({'success': False, 'error': 'لا توجد عملية للتراجع عنها'})
            
        undo_data = undo_info.get('data', [])
        if not undo_data:
            return JsonResponse({'success': False, 'error': 'بيانات التراجع فارغة'})
            
        with transaction.atomic():
            # أولاً: نحصل على جميع معرفات المنتجات المتأثرة
            product_ids = [item['product_id'] for item in undo_data]
            
            # نحصل على المنتجات الحالية
            products_map = {p.id: p for p in Product.objects.filter(id__in=product_ids)}
            
            # إعادة المنتجات لمواقعها الأصلية
            for item in undo_data:
                product = products_map.get(item['product_id'])
                if product:
                    product.location_id = item['location_id']
                    product.save(update_fields=['location'])
            
            # مسح بيانات التراجع
            del request.session['last_compaction_undo']
            request.session.modified = True
            
        type_str = "الصف" if undo_info['type'] == 'row' else "العمود"
        return JsonResponse({'success': True, 'message': f'تم التراجع عن ترتيب {type_str} {undo_info["id"]} بنجاح'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ أثناء التراجع: {str(e)}'})


@admin_required
def secure_backup_login(request):
    """تسجيل الدخول للصندوق الأسود"""
    if request.method == 'POST':
        password = request.POST.get('password')
        # كلمة مرور خاصة للصندوق الأسود
        if password == 'secure999':
            request.session['secure_backup_access'] = True
            return redirect('inventory_app:secure_backup_dashboard')
        else:
            return render(request, 'inventory_app/secure_backup_login.html', {'error': 'كلمة المرور غير صحيحة'})
    
    return render(request, 'inventory_app/secure_backup_login.html')


@admin_required
def secure_backup_dashboard(request):
    """لوحة تحكم السجل الآمن (الصندوق الأسود)"""
    # التحقق من صلاحية الوصول
    if not request.session.get('secure_backup_access'):
        return redirect('inventory_app:secure_backup_login')

    # Filters
    q = request.GET.get('q', '')
    table = request.GET.get('table', '')
    action = request.GET.get('action', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    queryset = SecureBackup.objects.all().order_by('-timestamp')

    if q:
        queryset = queryset.filter(db_models.Q(id__icontains=q) | db_models.Q(hash_signature__icontains=q))
    if table:
        queryset = queryset.filter(table_name=table)
    if action:
        queryset = queryset.filter(action=action)
    if date_from:
        queryset = queryset.filter(timestamp__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(timestamp__date__lte=date_to)

    # Stats
    stats = {
        'total': SecureBackup.objects.count(),
        'today': SecureBackup.objects.filter(timestamp__date=timezone.now().date()).count(),
        'deleted_items': SecureBackup.objects.filter(action='delete').count()
    }

    # Pagination
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    tables = SecureBackup.objects.values_list('table_name', flat=True).distinct()

    context = {
        'backups': page_obj,
        'stats': stats,
        'tables': tables,
        'current_filters': {
            'q': q, 'table': table, 'action': action,
            'date_from': date_from, 'date_to': date_to
        }
    }
    return render(request, 'inventory_app/secure_backup.html', context)


@admin_required
def get_secure_backup_detail(request, backup_id):
    # التحقق من صلاحية الوصول
    if not request.session.get('secure_backup_access'):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    backup = get_object_or_404(SecureBackup, id=backup_id)
    return JsonResponse({
        'id': backup.id,
        'table': backup.table_name,
        'record_id': backup.record_id,
        'action': backup.get_action_display(),
        'timestamp': backup.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        'data': backup.backup_data,
        'hash': backup.hash_signature
    })


@admin_required
def export_secure_backup(request):
    """تصدير سجلات الصندوق الأسود (SecureBackup)"""
    # التحقق من صلاحية الوصول
    if not request.session.get('secure_backup_access'):
        return redirect('inventory_app:secure_backup_login')
        
    try:
        # تصدير كل السجلات
        backups = SecureBackup.objects.all().order_by('timestamp')
        data = json.loads(serializers.serialize('json', backups))
        
        # إضافة معلومات وصفية
        export_data = {
            'meta': {
                'type': 'secure_backup_export',
                'date': datetime.now().isoformat(),
                'count': len(data),
                'description': 'تصدير كامل لسجلات الصندوق الأسود'
            },
            'records': data
        }
        
        json_data = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        filename = f'secure_backup_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)



@admin_required
def data_quality(request):
    """صفحة فحص جودة البيانات"""
    # 1. المنتجات بدون موقع
    products_no_location = Product.objects.filter(location__isnull=True).order_by('product_number')
    
    # 2. المنتجات أقل من 12 حبة
    products_low_qty = Product.objects.filter(quantity__lt=12).order_by('product_number')
    
    # 3. المنتجات بدون صور
    products_no_image = Product.objects.filter(db_models.Q(image__exact='') | db_models.Q(image__isnull=True)).order_by('product_number')
    
    # 4. الحاويات الفارغة
    empty_containers = Container.objects.annotate(prod_count=db_models.Count('product')).filter(prod_count=0)
    
    context = {
        'products_no_location': products_no_location,
        'products_low_qty': products_low_qty,
        'products_no_image': products_no_image,
        'empty_containers': empty_containers,
    }
    return render(request, 'inventory_app/data_quality.html', context)


@admin_required
def inventory_insights(request):
    """صفحة توصيات المخزون"""
    try:
        # تحديد حد انخفاض المخزون (تم التعديل إلى 24)
        low_stock_limit = 24
        
        # المنتجات منخفضة المخزون بشكل عام (أقل من أو يساوي الحد المحدد)
        general_low_stock = Product.objects.filter(quantity__lte=low_stock_limit).order_by('quantity')
        
        counts = {
            'low_general': general_low_stock.count()
        }
        
        policy = {
            'low_stock_limit': low_stock_limit
        }

        context = {
            'general_low_stock': general_low_stock,
            'counts': counts,
            'policy': policy
        }
        return render(request, 'inventory_app/inventory_insights.html', context)
    except Exception as e:
        import traceback
        print(f"Error in inventory_insights: {e}")
        traceback.print_exc()
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@require_http_methods(["GET"])
def get_warehouse_grid(request):
    """الحصول على شبكة المستودع"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    # Optimize query with prefetch_related
    locations = Location.objects.filter(warehouse=warehouse).prefetch_related('products')
    grid_data = {}
    
    for location in locations:
        key = f"{location.row},{location.column}"
        # Use prefetched products
        products = list(location.products.all())
        grid_data[key] = {
            'row': location.row,
            'column': location.column,
            'notes': location.notes,
            'is_active': location.is_active,
            'has_products': len(products) > 0,
            'products': [p.product_number for p in products],
        }
    
    return JsonResponse({
        'rows': warehouse.rows_count,
        'columns': warehouse.columns_count,
        'grid': grid_data
    })


@require_http_methods(["POST"])
@csrf_exempt
def add_row(request):
    """إضافة صف/صفوف جديدة"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الصفوف المراد إضافتها
        
        if count < 1 or count > 50:
            return JsonResponse({'error': 'العدد يجب أن يكون بين 1 و 50'}, status=400)
        
        with transaction.atomic():
            rows_added = 0
            for _ in range(count):
                warehouse.rows_count += 1
                warehouse.save()
                
                for col in range(1, warehouse.columns_count + 1):
                    Location.objects.create(
                        warehouse=warehouse,
                        row=warehouse.rows_count,
                        column=col
                    )
                rows_added += 1
            
            return JsonResponse({
                'success': True,
                'new_rows_count': warehouse.rows_count,
                'rows_added': rows_added
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["POST"])
@csrf_exempt
def add_column(request):
    """إضافة عمود/أعمدة جديدة"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الأعمدة المراد إضافتها
        
        if count < 1 or count > 50:
            return JsonResponse({'error': 'العدد يجب أن يكون بين 1 و 50'}, status=400)
        
        with transaction.atomic():
            columns_added = 0
            for _ in range(count):
                warehouse.columns_count += 1
                warehouse.save()
                
                for row in range(1, warehouse.rows_count + 1):
                    Location.objects.create(
                        warehouse=warehouse,
                        row=row,
                        column=warehouse.columns_count
                    )
                columns_added += 1
            
            return JsonResponse({
                'success': True,
                'new_columns_count': warehouse.columns_count,
                'columns_added': columns_added
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["POST"])
@csrf_exempt
def delete_row(request):
    """حذف صف/صفوف"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الصفوف المراد حذفها
        
        if count < 1 or count > warehouse.rows_count:
            return JsonResponse({'error': f'العدد يجب أن يكون بين 1 و {warehouse.rows_count}'}, status=400)
        
        with transaction.atomic():
            rows_deleted = 0
            for _ in range(count):
                if warehouse.rows_count > 0:
                    # حذف آخر صف
                    Location.objects.filter(warehouse=warehouse, row=warehouse.rows_count).delete()
                    warehouse.rows_count -= 1
                    warehouse.save()
                    rows_deleted += 1
                else:
                    break
            
            return JsonResponse({
                'success': True,
                'new_rows_count': warehouse.rows_count,
                'rows_deleted': rows_deleted
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["POST"])
@csrf_exempt
def delete_column(request):
    """حذف عمود/أعمدة"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الأعمدة المراد حذفها
        
        if count < 1 or count > warehouse.columns_count:
            return JsonResponse({'error': f'العدد يجب أن يكون بين 1 و {warehouse.columns_count}'}, status=400)
        
        with transaction.atomic():
            columns_deleted = 0
            for _ in range(count):
                if warehouse.columns_count > 0:
                    # حذف آخر عمود
                    Location.objects.filter(warehouse=warehouse, column=warehouse.columns_count).delete()
                    warehouse.columns_count -= 1
                    warehouse.save()
                    columns_deleted += 1
                else:
                    break
            
            return JsonResponse({
                'success': True,
                'new_columns_count': warehouse.columns_count,
                'columns_deleted': columns_deleted
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def warehouse_dashboard(request):
    """لوحة تحكم المستودع"""
    # Use select_related for related data
    warehouse = Warehouse.objects.first()
    products_count = Product.objects.count()
    locations_count = Location.objects.count()
    total_capacity = warehouse.rows_count * warehouse.columns_count if warehouse else 0
    
    # إحصائيات الحاويات
    containers_count = Container.objects.count()
    products_with_containers = Product.objects.filter(container__isnull=False).count()
    products_without_containers = Product.objects.filter(container__isnull=True).count()
    
    # إحصائيات الطلبات والمرتجعات
    orders_count = Order.objects.count()
    returns_count = ProductReturn.objects.count()
    # لا يوجد حقل status في ProductReturn، لذا نستخدم العدد الإجمالي
    pending_returns = 0  # يمكن إضافة حقل status لاحقاً إذا لزم الأمر
    
    # المنتجات النافذة (كمية = 0)
    out_of_stock_products = Product.objects.filter(quantity=0).count()
    
    # إجمالي الكمية في المخزون
    total_quantity = Product.objects.aggregate(
        total=db_models.Sum('quantity')
    )['total'] or 0
    
    # التقارير الذكية الأخيرة
    latest_ai_reports = []  # AIInsightLog.objects.all()[:5]
    
    return render(request, 'inventory_app/dashboard.html', {
        'warehouse': warehouse,
        'products_count': products_count,
        'locations_count': locations_count,
        'total_capacity': total_capacity,
        'containers_count': containers_count,
        'products_with_containers': products_with_containers,
        'products_without_containers': products_without_containers,
        'orders_count': orders_count,
        'returns_count': returns_count,
        'pending_returns': pending_returns,
        'out_of_stock_products': out_of_stock_products,
        'total_quantity': total_quantity,
        'latest_ai_reports': latest_ai_reports,
    })


def products_list(request):
    """قائمة جميع المنتجات"""
    from django.core.paginator import Paginator
    
    # Optimize query with select_related to avoid N+1 queries
    products = Product.objects.select_related('location', 'container').all().order_by('product_number')
    
    # احسب العدد الإجمالي قبل أي فلترة
    total_count = products.count()
    
    # فلترة حسب الحاوية (إذا تم تحديدها)
    container_id = request.GET.get('container', '')
    selected_container = None
    if container_id:
        try:
            selected_container = Container.objects.get(id=container_id)
            products = products.filter(container=selected_container)
        except (Container.DoesNotExist, ValueError):
            pass
    
    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            product_number__icontains=search
        ) | products.filter(
            name__icontains=search
        )
    
    # احسب العدد بعد الفلترة (إذا كان هناك بحث أو حاوية)
    filtered_count = products.count() if (search or container_id) else total_count
    
    # عرض جميع المنتجات بدون ترقيم صفحات بشكل افتراضي
    show_all = True
    page_obj = products
    page_obj.has_other_pages = False
    page_obj.number = 1
    page_obj.paginator = type('obj', (object,), {'num_pages': 1})()
    page_obj.start_index = lambda: 1
    page_obj.end_index = lambda: products.count()
    
    # جلب جميع الحاويات للقائمة المنسدلة
    containers = Container.objects.all().order_by('name')
    
    return render(request, 'inventory_app/products_list.html', {
        'products': page_obj,
        'page_obj': page_obj,
        'search': search,
        'total_count': total_count,
        'containers': containers,
        'filtered_count': filtered_count,
        'show_all': show_all,
        'selected_container': selected_container
    })


def product_detail(request, product_id):
    """تفاصيل منتج"""
    product = get_object_or_404(Product, id=product_id)
    first_log = product.audit_logs.order_by('created_at').first()
    original_quantity_ref = None
    if first_log and first_log.quantity_after is not None:
        original_quantity_ref = first_log.quantity_after
    else:
        original_quantity_ref = product.quantity
    return render(request, 'inventory_app/product_detail.html', {'product': product, 'original_quantity_ref': original_quantity_ref})


def product_add(request):
    """إضافة منتج جديد"""
    if request.method == 'POST':
        try:
            # إنشاء المنتج
            price_val = request.POST.get('price', '')
            if price_val:
                price_val = float(price_val)
            else:
                price_val = None

            product = Product.objects.create(
                product_number=request.POST.get('product_number'),
                name=request.POST.get('name', ''),
                # category=request.POST.get('category', ''), # Removed
                description=request.POST.get('description', ''),
                quantity=int(request.POST.get('quantity', 0) or 0),
                price=price_val
            )
            
            # رفع الصورة إذا تم اختيارها
            if 'image' in request.FILES:
                product.image = request.FILES['image']
                product.save()
            
            # تسجيل العملية في السجل
            AuditLog.objects.create(
                action='added',
                product=product,
                product_number=product.product_number,
                quantity_before=0,
                quantity_after=product.quantity,
                quantity_change=product.quantity,
                notes=f'تم إضافة منتج جديد: {product.name}',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            messages.success(request, 'تم إضافة المنتج بنجاح')
            return redirect('inventory_app:product_detail', product_id=product.id)
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    return render(request, 'inventory_app/product_add.html')


def product_edit(request, product_id):
    """تعديل منتج"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            # حفظ القيم القديمة
            old_product_number = product.product_number
            old_name = product.name
            # old_category = product.category  # Removed
            old_quantity = product.quantity
            old_price = product.price
            
            # التعديل
            new_product_number = request.POST.get('product_number')
            new_name = request.POST.get('name', '')
            # new_category = request.POST.get('category', '')  # Removed
            new_quantity = int(request.POST.get('quantity', 0) or 0)
            
            price_val = request.POST.get('price', '')
            if price_val:
                new_price = float(price_val)
            else:
                new_price = None
            
            product.product_number = new_product_number
            product.name = new_name
            # product.category = new_category  # Removed
            product.description = request.POST.get('description', '')
            product.quantity = new_quantity
            product.price = new_price
            
            # حذف الصورة إذا تم تحديد الخيار
            if request.POST.get('delete_image') == '1' and product.image:
                product.image.delete()
                product.image = None
            
            # رفع صورة جديدة إذا تم اختيارها
            if 'image' in request.FILES:
                # حذف الصورة القديمة إذا كانت موجودة
                if product.image:
                    product.image.delete()
                product.image = request.FILES['image']
            
            product.save()
            
            # تسجيل التغييرات في السجل
            changes = []
            if old_product_number != new_product_number:
                changes.append(f'رقم المنتج: {old_product_number} → {new_product_number}')
            if old_name != new_name:
                changes.append(f'الاسم: {old_name} → {new_name}')
            # if old_category != new_category:
            #     changes.append(f'الفئة: {old_category} → {new_category}')
            if old_quantity != new_quantity:
                changes.append(f'الكمية: {old_quantity} → {new_quantity}')
            if old_price != new_price:
                changes.append(f'السعر: {old_price} → {new_price}')
            
            if changes:
                AuditLog.objects.create(
                    action='updated',
                    product=product,
                    product_number=product.product_number,
                    quantity_before=old_quantity,
                    quantity_after=new_quantity,
                    quantity_change=new_quantity - old_quantity,
                    notes='تغييرات: ' + ' | '.join(changes),
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
            
            messages.success(request, 'تم تعديل المنتج بنجاح')
            return redirect('inventory_app:product_detail', product_id=product.id)
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    return render(request, 'inventory_app/product_edit.html', {'product': product})


@csrf_exempt
def product_delete(request, product_id):
    """حذف منتج"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            product_number = product.product_number
            quantity = product.quantity
            name = product.name
            
            # تسجيل العملية قبل الحذف مع لقطة كاملة
            snapshot = {
                'product_number': product_number,
                'name': product.name,
                'category': product.category,
                'description': product.description,
                'container': product.container.name if product.container else None,
                'location': product.location.full_location if product.location else None,
                'quantity': product.quantity,
                'min_stock_threshold': product.min_stock_threshold,
                'store_quantity': product.store_quantity,
                'warehouse_quantity': product.warehouse_quantity,
                'barcode': product.barcode,
                'image_url': product.image_url,
                'price': str(product.price) if product.price is not None else None,
            }
            AuditLog.objects.create(
                action='deleted',
                product=product,
                product_number=product_number,
                quantity_before=quantity,
                quantity_after=0,
                quantity_change=-quantity,
                notes=f'تم حذف المنتج: {name}',
                product_snapshot=snapshot,
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            product.delete()
            
            # إرجاع JSON للطلبات AJAX
            if request.content_type == 'application/json':
                return JsonResponse({'success': True, 'message': f'تم حذف المنتج {product_number}'}, json_dumps_params={'ensure_ascii': False})
            
            messages.success(request, f'تم حذف المنتج {product_number}')
            return redirect('inventory_app:products_list')
        except Exception as e:
            if request.content_type == 'application/json':
                return JsonResponse({'success': False, 'error': str(e)}, json_dumps_params={'ensure_ascii': False})
            messages.error(request, f'خطأ في حذف المنتج: {str(e)}')
            return redirect('inventory_app:product_detail', product_id=product.id)
    
    return render(request, 'inventory_app/product_delete.html', {'product': product})


@login_required
def restore_product(request, log_id):
    """استعادة منتج محذوف من السجلات"""
    if not request.user.user_profile.is_admin():
        messages.error(request, 'غير مصرح لك بالقيام بهذا الإجراء')
        return redirect('inventory_app:audit_logs')
        
    log_entry = get_object_or_404(AuditLog, id=log_id, action='deleted')
    snapshot = log_entry.product_snapshot
    
    if not snapshot:
        messages.error(request, 'لا توجد نسخة احتياطية للمنتج')
        return redirect('inventory_app:audit_logs')
        
    # التحقق من وجود منتج بنفس الرقم
    if Product.objects.filter(product_number=snapshot.get('product_number')).exists():
        messages.error(request, f'منتج بنفس الرقم {snapshot.get("product_number")} موجود بالفعل')
        return redirect('inventory_app:audit_logs')

    try:
        # استعادة الحاوية
        container = None
        if snapshot.get('container'):
            container = Container.objects.filter(name=snapshot.get('container')).first()
            
        # إنشاء المنتج
        product = Product.objects.create(
            product_number=snapshot.get('product_number'),
            name=snapshot.get('name'),
            category=snapshot.get('category'),
            description=snapshot.get('description'),
            quantity=snapshot.get('quantity', 0),
            container=container,
            # الموقع لا يمكن استعادته بدقة لأنه قد يكون مشغولاً، لذا نتركه فارغاً
            location=None,
            min_stock_threshold=snapshot.get('min_stock_threshold', 0),
            price=snapshot.get('price'),
            store_quantity=snapshot.get('store_quantity', 0),
            warehouse_quantity=snapshot.get('warehouse_quantity', 0),
            barcode=snapshot.get('barcode'),
            image_url=snapshot.get('image_url'),
            colors=snapshot.get('colors', list())
        )
        
        # تسجيل عملية الاستعادة
        AuditLog.objects.create(
            action='added',
            product=product,
            product_number=product.product_number,
            quantity_before=0,
            quantity_after=product.quantity,
            quantity_change=product.quantity,
            notes=f'تم استعادة المنتج: {product.name} من السجل',
            user=request.user.username
        )
        
        messages.success(request, f'تم استعادة المنتج {product.name} بنجاح')
    except Exception as e:
        messages.error(request, f'حدث خطأ أثناء الاستعادة: {str(e)}')
        
    return redirect('inventory_app:audit_logs')


@csrf_exempt
def delete_products_bulk(request):
    """حذف منتجات متعددة دفعة واحدة"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_ids = data.get('product_ids', [])
            
            if not product_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'لم يتم تحديد أي منتجات'
                }, json_dumps_params={'ensure_ascii': False})
            
            deleted_count = 0
            deleted_products = []
            
            with transaction.atomic():
                products = Product.objects.filter(id__in=product_ids)
                
                for product in products:
                    product_number = product.product_number
                    quantity = product.quantity
                    name = product.name
                    
                    # تسجيل العملية قبل الحذف مع لقطة كاملة
                    snapshot = {
                        'product_number': product_number,
                        'name': product.name,
                        'category': product.category,
                        'description': product.description,
                        'container': product.container.name if product.container else None,
                        'location': product.location.full_location if product.location else None,
                        'quantity': product.quantity,
                        'min_stock_threshold': product.min_stock_threshold,
                        'store_quantity': product.store_quantity,
                        'warehouse_quantity': product.warehouse_quantity,
                        'barcode': product.barcode,
                        'image_url': product.image_url,
                        'price': str(product.price) if product.price is not None else None,
                        'colors': product.colors,
                    }
                    AuditLog.objects.create(
                        action='deleted',
                        product=product,
                        product_number=product_number,
                        quantity_before=quantity,
                        quantity_after=0,
                        quantity_change=-quantity,
                        notes=f'حذف جماعي: {name}',
                        product_snapshot=snapshot,
                        user=request.user.username if request.user.is_authenticated else 'Guest'
                    )
                    
                    deleted_products.append(product_number)
                
                # حذف المنتجات
                products.delete()
                deleted_count = len(deleted_products)
            
            return JsonResponse({
                'success': True,
                'message': f'تم حذف {deleted_count} منتج بنجاح',
                'deleted_count': deleted_count,
                'deleted_products': deleted_products[:10]  # أول 10 فقط
            }, json_dumps_params={'ensure_ascii': False})
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, json_dumps_params={'ensure_ascii': False})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)


@csrf_exempt
def move_product_with_shift(request, product_id):
    """نقل منتج لموقع معين وإعادة ترتيب باقي المنتجات في نفس العمود تلقائياً"""
    try:
        product = get_object_or_404(Product, id=product_id)
        data = json.loads(request.body)
        
        new_location_str = data.get('new_location', '')
        if not new_location_str:
            return JsonResponse({'success': False, 'error': 'الموقع الجديد مطلوب'}, json_dumps_params={'ensure_ascii': False})
        
        # تحليل الموقع الجديد (مثال: R15C4)
        import re
        match = re.match(r'R(\d+)C(\d+)', new_location_str)
        if not match:
            return JsonResponse({'success': False, 'error': 'تنسيق الموقع غير صحيح'}, json_dumps_params={'ensure_ascii': False})
        
        new_row = int(match.group(1))
        new_column = int(match.group(2))
        
        warehouse = Warehouse.objects.first()
        if not warehouse:
            return JsonResponse({'success': False, 'error': 'لا يوجد مستودع'}, json_dumps_params={'ensure_ascii': False})
        
        # التحقق من أن العمود الجديد يحتوي على عدد صفوف كافي
        if new_row > warehouse.rows_count:
            # إضافة صفوف إضافية
            for row in range(warehouse.rows_count + 1, new_row + 1):
                warehouse.rows_count += 1
                warehouse.save()
                
                # إنشاء المواقع للعمود الجديد
                for col in range(1, warehouse.columns_count + 1):
                    Location.objects.get_or_create(
                        warehouse=warehouse,
                        row=row,
                        column=col,
                        defaults={'is_active': True}
                    )
        
        # التحقق من المساحة المتاحة في العمود C قبل النقل
        old_location = product.location
        old_row = old_location.row if old_location else None
        old_column = old_location.column if old_location else None
        
        # حساب عدد المنتجات الموجودة في العمود C (باستثناء المنتج الحالي إذا كان في نفس العمود)
        products_in_column = Product.objects.filter(
            location__warehouse=warehouse,
            location__column=new_column
        )
        
        # إذا كان المنتج ينقل من نفس العمود، لا نحسبه في العدد
        if old_column == new_column:
            products_in_column = products_in_column.exclude(id=product.id)
        
        products_count_in_column = products_in_column.count()
        
        # التحقق من وجود منتج في الموقع الجديد
        new_location = Location.objects.get(warehouse=warehouse, row=new_row, column=new_column)
        existing_product = Product.objects.filter(location=new_location).exclude(id=product.id).first()
        
        # إذا كان هناك منتج في الموقع الجديد، نحتاج لمساحة إضافية
        if existing_product:
            # حساب آخر منتج في العمود (أعلى صف مشغول)
            last_product = products_in_column.order_by('-location__row').first()
            if last_product:
                last_occupied_row = last_product.location.row
            else:
                last_occupied_row = new_row - 1
            
            # عدد المنتجات التي تحتاج للانتقال لأسفل (من الموقع الجديد فما تحته)
            products_to_shift = Product.objects.filter(
                location__warehouse=warehouse,
                location__column=new_column,
                location__row__gte=new_row
            ).exclude(id=product.id).count()
            
            # آخر صف مطلوب بعد النقل
            last_row_needed = new_row + products_to_shift
            
            # إذا كان آخر صف مطلوب أكبر من عدد الصفوف المتاحة، العمود ممتلئ
            if last_row_needed > warehouse.rows_count:
                # حساب عدد الصفوف المتاحة
                available_rows = warehouse.rows_count
                return JsonResponse({
                    'success': False, 
                    'error': f'العمود C{new_column} ممتلئ! لا توجد مساحة كافية. عدد الصفوف المتاحة: {available_rows} صف'
                }, json_dumps_params={'ensure_ascii': False})
        
        # إذا كان العمود ممتلئ بالكامل (عدد المنتجات = عدد الصفوف) ولا يوجد مساحة للنقل
        if products_count_in_column >= warehouse.rows_count and not existing_product:
            # إذا كان المنتج ينقل لنفس العمود، لا مشكلة
            if old_column != new_column:
                return JsonResponse({
                    'success': False, 
                    'error': f'العمود C{new_column} ممتلئ بالكامل! عدد المنتجات: {products_count_in_column}، عدد الصفوف المتاحة: {warehouse.rows_count}'
                }, json_dumps_params={'ensure_ascii': False})
        
        with transaction.atomic():
            old_location_str = old_location.full_location if old_location else 'بدون موقع'
            new_location_str = new_location.full_location
            
            # إذا كان الموقع الجديد مشغول بمنتج آخر، يجب إعادة ترتيب المنتجات في نفس العمود
            if existing_product:
                # جلب جميع المنتجات التي في نفس العمود الجديد وتبدأ من الموقع الجديد أو تحته
                # ترتيبها من الأعلى للأسفل
                products_to_shift_down = Product.objects.filter(
                    location__warehouse=warehouse,
                    location__column=new_column,
                    location__row__gte=new_row
                ).exclude(id=product.id).select_related('location').order_by('location__row')
                
                # حساب عدد المنتجات التي تحتاج للانتقال لأسفل
                products_count = products_to_shift_down.count()
                
                # إذا كان هناك منتجات تحتاج للانتقال، نحتاج للتأكد من وجود صفوف كافية
                if products_count > 0:
                    # حساب آخر صف مطلوب
                    last_row_needed = new_row + products_count
                    
                    # إذا كان آخر صف مطلوب أكبر من عدد الصفوف المتاحة، إضافة صفوف جديدة
                    if last_row_needed > warehouse.rows_count:
                        rows_to_add = last_row_needed - warehouse.rows_count
                        for row in range(warehouse.rows_count + 1, last_row_needed + 1):
                            warehouse.rows_count += 1
                            warehouse.save()
                            
                            # إنشاء المواقع للصفوف الجديدة
                            for col in range(1, warehouse.columns_count + 1):
                                Location.objects.get_or_create(
                                    warehouse=warehouse,
                                    row=row,
                                    column=col,
                                    defaults={'is_active': True}
                                )
                
                # نقل المنتجات من الأسفل للأعلى (من آخر منتج لأول منتج) لتجنب التعارض
                products_list = list(products_to_shift_down)
                products_list.reverse()  # نبدأ من آخر منتج
                
                for prod in products_list:
                    old_loc = prod.location
                    # نقل المنتج صف واحد لأسفل
                    new_row_for_prod = old_loc.row + 1
                    
                    new_loc = Location.objects.get(
                        warehouse=warehouse,
                        row=new_row_for_prod,
                        column=new_column
                    )
                    
                    prod.location = new_loc
                    prod.save()
                    
                    AuditLog.objects.create(
                        action='location_assigned',
                        product=prod,
                        product_number=prod.product_number,
                        quantity_before=prod.quantity,
                        quantity_after=prod.quantity,
                        quantity_change=0,
                        notes=f'إعادة ترتيب تلقائي: {old_loc.full_location} → {new_loc.full_location}',
                        user=request.user.username if request.user.is_authenticated else 'Guest'
                    )
            
            # الآن نقل المنتج إلى الموقع الجديد
            product.location = new_location
            product.save()
            
            # تسجيل العملية للمنتج المنقول
            AuditLog.objects.create(
                action='location_assigned',
                product=product,
                product_number=product.product_number,
                quantity_before=product.quantity,
                quantity_after=product.quantity,
                quantity_change=0,
                notes=f'نقل مع إعادة ترتيب: {old_location_str} → {new_location_str}',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            # إذا كان النقل في نفس العمود، نحتاج لإعادة ترتيب المنتجات في الموقع القديم
            if old_location and old_column == new_column:
                # سيناريو 1: نقل المنتج لأسفل في نفس العمود (من R10 إلى R15)
                if old_row < new_row:
                    # جلب المنتجات التي كانت بين الموقع القديم والجديد
                    products_to_shift_up = Product.objects.filter(
                        location__warehouse=warehouse,
                        location__column=old_column,
                        location__row__gt=old_row,
                        location__row__lt=new_row  # أقل من الموقع الجديد (لأن المنتج الآن في الموقع الجديد)
                    ).select_related('location').order_by('location__row')
                    
                    # نقل هذه المنتجات صف واحد لأعلى
                    for prod in products_to_shift_up:
                        old_loc = prod.location
                        # نقل المنتج لصف واحد أقل
                        if old_loc.row > 1:
                            new_loc = Location.objects.get(
                                warehouse=warehouse,
                                row=old_loc.row - 1,
                                column=old_loc.column
                            )
                            prod.location = new_loc
                            prod.save()
                            
                            AuditLog.objects.create(
                                action='location_assigned',
                                product=prod,
                                product_number=prod.product_number,
                                quantity_before=prod.quantity,
                                quantity_after=prod.quantity,
                                quantity_change=0,
                                notes=f'إعادة ترتيب تلقائي: {old_loc.full_location} → {new_loc.full_location}',
                                user=request.user.username if request.user.is_authenticated else 'Guest'
                            )
                
                # سيناريو 2: نقل المنتج لأعلى في نفس العمود (من R15 إلى R10)
                elif old_row > new_row:
                    # جلب المنتجات التي كانت بين الموقع الجديد والقديم
                    # (لا حاجة لهذا لأننا نتعامل مع الموقع الجديد أعلاه)
                    pass
            
            return JsonResponse({
                'success': True,
                'message': f'تم نقل المنتج {product.product_number} بنجاح من {old_location_str} إلى {new_location_str}'
            }, json_dumps_params={'ensure_ascii': False})
            
    except Location.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الموقع المطلوب غير موجود'}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, json_dumps_params={'ensure_ascii': False})


def assign_location_to_product(request, product_id):
    """ربط منتج بموقع واحد فقط"""
    # Optimize with select_related
    product = get_object_or_404(Product.objects.select_related('location'), id=product_id)
    warehouse = Warehouse.objects.first()
    
    # إنشاء جميع المواقع المفقودة
    if warehouse:
        locations_created = 0
        with transaction.atomic():
            for row in range(1, warehouse.rows_count + 1):
                for col in range(1, warehouse.columns_count + 1):
                    location, created = Location.objects.get_or_create(
                        warehouse=warehouse,
                        row=row,
                        column=col,
                        defaults={
                            'is_active': True
                        }
                    )
                    if created:
                        locations_created += 1
    
    if request.method == 'POST':
        location_id = request.POST.get('location')
        if location_id:
            new_location = Location.objects.get(id=location_id)
            old_location = product.location
            
            # التحقق من أن الموقع غير مشغول بمنتج آخر - use exists() for efficiency
            if new_location.products.exclude(id=product.id).exists():
                messages.error(request, 'هذا الموقع مشغول بمنتج آخر! اختر موقعاً آخر')
            else:
                product.location = new_location
                product.save()
                
                # تسجيل العملية
                old_location_str = old_location.full_location if old_location else 'بدون موقع'
                new_location_str = new_location.full_location
                
                AuditLog.objects.create(
                    action='location_assigned',
                    product=product,
                    product_number=product.product_number,
                    quantity_before=product.quantity,
                    quantity_after=product.quantity,
                    quantity_change=0,
                    notes=f'تغيير الموقع: {old_location_str} → {new_location_str}',
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
                
                messages.success(request, f'تم ربط المنتج بالموقع {new_location_str} بنجاح')
        else:
            # إلغاء الربط
            if product.location:
                old_location_str = product.location.full_location
                product.location = None
                product.save()
                
                AuditLog.objects.create(
                    action='location_removed',
                    product=product,
                    product_number=product.product_number,
                    quantity_before=product.quantity,
                    quantity_after=product.quantity,
                    quantity_change=0,
                    notes=f'إلغاء ربط الموقع: {old_location_str}',
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
                
                messages.success(request, 'تم إلغاء ربط الموقع')
        
        return redirect('inventory_app:product_detail', product_id=product.id)
    
    # Optimize query
    all_locations = Location.objects.filter(warehouse=warehouse).select_related('warehouse').order_by('row', 'column')
    
    # تحديد الأماكن الشاغرة والفارغة - optimize with values_list
    occupied_locations = set(
        Location.objects.filter(products__isnull=False)
        .values_list('id', flat=True)
    )
    
    return render(request, 'inventory_app/assign_location.html', {
        'product': product,
        'all_locations': all_locations,
        'current_location': product.location,
        'warehouse': warehouse,
        'occupied_locations': occupied_locations
    })


def warehouses_list(request):
    """قائمة المستودعات"""
    warehouses = Warehouse.objects.all()
    return render(request, 'inventory_app/warehouses_list.html', {'warehouses': warehouses})


def warehouse_detail(request, warehouse_id):
    """تفاصيل مستودع"""
    warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    locations_count = Location.objects.filter(warehouse=warehouse).count()
    return render(request, 'inventory_app/warehouse_detail.html', {
        'warehouse': warehouse,
        'locations_count': locations_count
    })


def locations_list(request):
    """قائمة جميع الأماكن"""
    warehouse = Warehouse.objects.first()
    
    if warehouse:
        # إنشاء جميع المواقع المفقودة
        locations_created = 0
        with transaction.atomic():
            for row in range(1, warehouse.rows_count + 1):
                for col in range(1, warehouse.columns_count + 1):
                    location, created = Location.objects.get_or_create(
                        warehouse=warehouse,
                        row=row,
                        column=col,
                        defaults={
                            'is_active': True
                        }
                    )
                    if created:
                        locations_created += 1
        
        if locations_created > 0:
            print(f'Created {locations_created} missing locations')
    
    # Optimize with prefetch_related and select_related
    # الحصول على جميع المواقع للشبكة (بدون pagination)
    all_locations = Location.objects.filter(warehouse=warehouse).select_related('warehouse').prefetch_related('products').order_by('row', 'column')
    
    # للحصول على paginated locations للقائمة (إن وجدت في المستقبل)
    search = request.GET.get('search', '')
    if search:
        all_locations = all_locations.filter(notes__icontains=search)
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(all_locations, 100)  # Show 100 locations per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'inventory_app/locations_list.html', {
        'locations': all_locations,  # تمرير جميع المواقع للشبكة
        'page_obj': page_obj,
        'warehouse': warehouse,
        'search': search
    })


@require_http_methods(["GET"])
def get_stats(request):
    """API للحصول على إحصائيات النظام"""
    try:
        from .models import Product, Location, Warehouse
        
        # إحصائيات المنتجات
        products_count = Product.objects.count()
        # المنتجات التي لها موقع (location ليست null)
        products_with_locations = Product.objects.filter(location__isnull=False).count()
        products_without_locations = products_count - products_with_locations
        
        # إحصائيات الأماكن
        locations_count = Location.objects.count()
        warehouse = Warehouse.objects.first()
        
        if warehouse:
            total_capacity = warehouse.rows_count * warehouse.columns_count
            # المواقع المشغولة (التي تحتوي على منتجات)
            # استخدام values_list للحصول على الـ IDs المميزة للأماكن المشغولة
            occupied_locations_ids = Product.objects.filter(location__isnull=False).values_list('location_id', flat=True).distinct()
            occupied_locations = len(occupied_locations_ids) if occupied_locations_ids else 0
            empty_locations = total_capacity - occupied_locations
        else:
            total_capacity = 0
            occupied_locations = 0
            empty_locations = 0
        
        low_stock_count = Product.objects.filter(quantity__lt=24, quantity__gt=0).count()
        out_of_stock_count = Product.objects.filter(quantity=0).count()

        # تم إزالة الحسابات الأخرى بناءً على طلب المستخدم (فقط المخزون المنخفض)
        reorder_count = 0
        overstock_count = 0
        watchlist_count = 0
        anomaly_count = 0

        return JsonResponse({
            # المنتجات
            'products_count': products_count,
            'products_with_locations': products_with_locations,
            'products_without_locations': products_without_locations,
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count,
            'reorder_count': 0,
            'overstock_count': 0,
            'watchlist_count': 0,
            'anomaly_count': 0,
            # الأماكن
            'locations_count': locations_count,
            'total_capacity': total_capacity,
            'occupied_locations': occupied_locations,
            'empty_locations': empty_locations,
            # معلومات المستودع
            'warehouse_rows': warehouse.rows_count if warehouse else 0,
            'warehouse_columns': warehouse.columns_count if warehouse else 0,
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def quick_search_products(request):
    """API للبحث السريع في المنتجات"""
    try:
        from .models import Product
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse([], safe=False, json_dumps_params={'ensure_ascii': False})
        
        # البحث في رقم المنتج واسمه
        products = Product.objects.filter(
            product_number__icontains=query
        ) | Product.objects.filter(
            name__icontains=query
        )
        
        products = products[:10]  # أول 10 نتائج
        
        results = []
        for product in products:
            results.append({
                'id': product.id,
                'product_number': product.product_number,
                'name': product.name,
                'location': product.location.full_location if product.location else None,
                'quantity': product.quantity
            })
        
        return JsonResponse(results, safe=False, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def quick_search_locations(request):
    """API للبحث السريع في الأماكن"""
    try:
        from .models import Location
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse([], safe=False, json_dumps_params={'ensure_ascii': False})
        
        # البحث في المواقع (R1C1, R2C3, etc.)
        locations = Location.objects.filter(
            row__icontains=query
        ) | Location.objects.filter(
            column__icontains=query
        )
        
        locations = locations[:10]  # أول 10 نتائج
        
        results = []
        for location in locations:
            has_product = location.products.exists()
            results.append({
                'id': location.id,
                'full_location': location.full_location,
                'row': location.row,
                'column': location.column,
                'has_product': has_product,
                'warehouse': location.warehouse.name if location.warehouse else ''
            })
        
        return JsonResponse(results, safe=False, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def audit_logs(request):
    """صفحة عرض سجلات العمليات"""
    from django.core.paginator import Paginator
    from django.db.models import Count

    # الاستعلام الأساسي + تحسينات
    base_qs = AuditLog.objects.select_related('product').all()

    search = request.GET.get('search', '')
    action_filter = request.GET.get('action', '')

    if search:
        base_qs = base_qs.filter(product_number__icontains=search)

    if action_filter:
        base_qs = base_qs.filter(action=action_filter)

    # إحصائيات كاملة قبل التقسيم لصفحات
    total_count = base_qs.count()
    counts_qs = base_qs.values('action').annotate(count=Count('id'))
    action_counts = {row['action']: row['count'] for row in counts_qs}

    # ترتيب لضمان تجميع منطقي داخل الصفحة (ثم الأحدث داخل كل نوع)
    ordered_qs = base_qs.order_by('action', '-created_at')

    # التقسيم لصفحات
    paginator = Paginator(ordered_qs, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    from django.utils import timezone
    from datetime import timedelta
    now = timezone.now()
    initial_period = request.GET.get('period', 'day')
    if initial_period not in ('day', 'week', 'month'):
        initial_period = 'day'
    last_24h = now - timedelta(hours=24)
    recent_qs = AuditLog.objects.filter(created_at__gte=last_24h)
    recent_counts_qs = recent_qs.values('action').annotate(count=Count('id'))
    recent_action_counts = {row['action']: row['count'] for row in recent_counts_qs}
    top_products_recent = list(
        recent_qs.values('product_number').annotate(count=Count('id')).order_by('-count')[:5]
    )
    daily_summary = {
        'total': recent_qs.count(),
        'actions': recent_action_counts,
        'top_products': top_products_recent,
    }

    seven_days_ago = now - timedelta(days=7)
    weekly_qs = AuditLog.objects.filter(created_at__gte=seven_days_ago)
    weekly_counts_qs = weekly_qs.values('action').annotate(count=Count('id'))
    weekly_action_counts = {row['action']: row['count'] for row in weekly_counts_qs}
    top_products_week = list(
        weekly_qs.values('product_number').annotate(count=Count('id')).order_by('-count')[:5]
    )
    weekly_summary = {
        'total': weekly_qs.count(),
        'actions': weekly_action_counts,
        'top_products': top_products_week,
    }

    thirty_days_ago = now - timedelta(days=30)
    monthly_qs = AuditLog.objects.filter(created_at__gte=thirty_days_ago)
    monthly_counts_qs = monthly_qs.values('action').annotate(count=Count('id'))
    monthly_action_counts = {row['action']: row['count'] for row in monthly_counts_qs}
    top_products_month = list(
        monthly_qs.values('product_number').annotate(count=Count('id')).order_by('-count')[:5]
    )
    monthly_summary = {
        'total': monthly_qs.count(),
        'actions': monthly_action_counts,
        'top_products': top_products_month,
    }

    return render(request, 'inventory_app/audit_logs.html', {
        'logs': page_obj,
        'page_obj': page_obj,
        'search': search,
        'action_filter': action_filter,
        'total_count': total_count,
        'action_counts': action_counts,
        'daily_summary': daily_summary,
        'weekly_summary': weekly_summary,
        'monthly_summary': monthly_summary,
        'initial_period': initial_period,
    })

def data_quality_report(request):
    products = Product.objects.select_related('location').all()
    import re
    from django.db.models import F, Q
    name_groups = {}
    for p in products:
        key = (p.name or '').strip().lower()
        if not key:
            continue
        name_groups.setdefault(key, []).append(p)
    duplicates_by_name = [
        {'key': k, 'items': v}
        for k, v in name_groups.items() if len(v) > 1
    ]

    barcode_groups = {}
    for p in products:
        key = (p.barcode or '').strip()
        if not key:
            continue
        barcode_groups.setdefault(key, []).append(p)
    duplicates_by_barcode = [
        {'key': k, 'items': v}
        for k, v in barcode_groups.items() if len(v) > 1
    ]

    number_groups = {}
    for p in products:
        raw = (p.product_number or '').strip()
        norm = re.sub(r'[^A-Za-z0-9]', '', raw).lower()
        if not norm:
            continue
        number_groups.setdefault(norm, []).append(p)
    near_duplicates_by_number = [
        {'key': k, 'items': v}
        for k, v in number_groups.items() if len(v) > 1
    ]

    missing_location_products = Product.objects.filter(location__isnull=True)

    invalid_carton_data = []
    # منطق الكراتين تم إيقافه
    
    return render(request, 'inventory_app/data_quality.html', {
        'duplicates_by_name': duplicates_by_name,
        'duplicates_by_barcode': duplicates_by_barcode,
        'near_duplicates_by_number': near_duplicates_by_number,
        'missing_location_products': missing_location_products,
        'invalid_carton_data': [],
        'counts': {
            'name': len(duplicates_by_name),
            'barcode': len(duplicates_by_barcode),
            'number': len(near_duplicates_by_number),
            'missing_location': missing_location_products.count(),
            'invalid_carton': 0,
        }
    })



@require_http_methods(["GET"])
def low_stock_products_api(request):
    try:
        limit = int(request.GET.get('limit', 10))
    except (ValueError, TypeError):
        limit = 10
    if limit < 1:
        limit = 10
    qs = Product.objects.filter(quantity__lt=24, quantity__gt=0).order_by('quantity')[:limit]
    data = []
    for p in qs:
        data.append({
            'id': p.id,
            'product_number': p.product_number,
            'name': p.name,
            'quantity': p.quantity,
        })
    return JsonResponse(data, safe=False, json_dumps_params={'ensure_ascii': False})




# ========== تصدير البيانات ==========

def export_products_pdf(request):
    """تصدير قائمة المنتجات إلى PDF احترافي مع دعم كامل للعربية باستخدام Playwright"""
    from django.http import HttpResponse
    from playwright.sync_api import sync_playwright
    from datetime import datetime
    import io
    
    try:
        # جلب المنتجات
        products = Product.objects.select_related('location').all().order_by('product_number')
        
        # تاريخ التقرير
        now = datetime.now()
        
        # إنشاء HTML محتوى
        html_content = f'''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', 'Arial', 'Tahoma', sans-serif;
            font-size: 10pt;
            direction: rtl;
            padding: 20px;
        }}
        
        .header {{
            text-align: center;
            color: #667eea;
            font-size: 28pt;
            font-weight: bold;
            margin-bottom: 20px;
        }}
        
        .info {{
            text-align: center;
            margin-bottom: 20px;
            font-size: 11pt;
            color: #374151;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 0 auto;
        }}
        
        th {{
            background-color: #667eea;
            color: white;
            padding: 12px 8px;
            text-align: right;
            font-weight: bold;
            border: 1px solid #555;
            font-size: 11pt;
        }}
        
        td {{
            padding: 8px;
            border: 1px solid #ddd;
            text-align: right;
            font-size: 9pt;
        }}
        
        tr:nth-child(even) {{
            background-color: #f8fafc;
        }}
        
        .summary {{
            margin-top: 20px;
            text-align: right;
            font-weight: bold;
            font-size: 12pt;
            color: #374151;
        }}
    </style>
</head>
<body>
    <div class="header">قائمة المنتجات</div>
    
    <div class="info">
        <strong>التاريخ:</strong> {now.strftime("%Y-%m-%d")} | 
        <strong>الوقت:</strong> {now.strftime("%H:%M")}
    </div>
    
    <table>
        <thead>
            <tr>
                <th>#</th>
                <th>رقم المنتج</th>
                <th>الاسم</th>
                <th>الفئة</th>
                <th>الكمية</th>
                <th>الموقع</th>
            </tr>
        </thead>
        <tbody>
'''
        
        # إضافة صفوف المنتجات
        for idx, product in enumerate(products, start=1):
            location = product.location.full_location if product.location else 'بدون موقع'
            category = product.category if product.category else '-'
            
            html_content += f'''
            <tr>
                <td>{idx}</td>
                <td>{product.product_number}</td>
                <td>{product.name}</td>
                <td>{category}</td>
                <td>{product.quantity}</td>
                <td>{location}</td>
            </tr>
            '''
        
        # إغلاق HTML
        html_content += f'''
        </tbody>
    </table>
    
    <div class="summary">
        إجمالي المنتجات: {products.count()}
    </div>
</body>
</html>
'''
        
        # إنشاء PDF باستخدام Playwright
        with sync_playwright() as p:
                    try:
                        browser = p.chromium.launch(headless=True)
                    except Exception as e:
                        # إذا فشل التشغيل، نحاول تثبيت المتصفح تلقائياً
                        print(f"Playwright launch failed: {e}. Attempting to install chromium...")
                        subprocess.run(["playwright", "install", "chromium"], check=True)
                        browser = p.chromium.launch(headless=True)

                    page = browser.new_page()
            page.set_content(html_content)
            
            pdf_bytes = page.pdf(
                format='A4',
                landscape=True,
                margin={'top': '1cm', 'right': '1cm', 'bottom': '1cm', 'left': '1cm'}
            )
            
            browser.close()
        
        # إرجاع الاستجابة
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="products_list.pdf"'
        
        return response
        
    except Exception as e:
        import traceback
        error_msg = f'خطأ في إنشاء PDF: {str(e)}\n{traceback.format_exc()}'
        return HttpResponse(error_msg, content_type='text/plain')

def export_order_pdf(request, order_id):
    from django.http import HttpResponse
    from playwright.sync_api import sync_playwright
    from datetime import datetime
    try:
        order = get_object_or_404(Order, id=order_id)
        
        # تحضير الصور
        product_numbers = [p.get('product_number') for p in order.products_data if p.get('product_number')]
        products = Product.objects.filter(product_number__in=product_numbers)
        products_map = {p.product_number: p for p in products}

        rows_html = ''
        for idx, item in enumerate(order.products_data, start=1):
            num = item.get('product_number')
            taken = item.get('quantity_taken')
            
            img_html = '-'
            price = 0
            if num in products_map:
                prod = products_map[num]
                price = float(prod.price) if prod.price is not None else 0
                if prod.image:
                    try:
                        import base64
                        with open(prod.image.path, "rb") as image_file:
                            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                            img_html = f'<img src="data:image/jpeg;base64,{encoded_string}" style="max-width: 50px; max-height: 50px; object-fit: contain;">'
                    except:
                        pass
                elif prod.image_url:
                    img_html = f'<img src="{prod.image_url}" style="max-width: 50px; max-height: 50px; object-fit: contain;">'

            rows_html += f'''
            <tr>
                <td>{idx}</td>
                <td style="text-align: center;"><strong>{num}</strong></td>
                <td style="text-align: center;">{img_html}</td>
                <td style="text-align: center;">{price}</td>
                <td style="text-align: center;">{taken}</td>
            </tr>
            '''

        html_content = f'''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', 'Arial', 'Tahoma', sans-serif; font-size: 10pt; direction: rtl; padding: 20px; }}
        .header-title {{ text-align: center; color: #1e293b; font-size: 16pt; font-weight: bold; margin-bottom: 15px; border-bottom: 2px solid #e2e8f0; padding-bottom: 5px; }}
        
        .info-grid {{ 
            display: grid; 
            grid-template-columns: repeat(2, 1fr); 
            gap: 10px; 
            margin-bottom: 20px; 
            background: #f8fafc;
            padding: 15px;
            border-radius: 8px;
            border: 1px solid #e2e8f0;
        }}
        
        .info-item {{ margin-bottom: 5px; }}
        .label {{ color: #64748b; font-weight: bold; font-size: 9pt; margin-bottom: 2px; display: block; }}
        .value {{ color: #0f172a; font-weight: bold; font-size: 11pt; }}
        
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        th {{ background: #1e293b; color: white; font-weight: bold; padding: 8px; border: 1px solid #1e293b; font-size: 10pt; }}
        td {{ padding: 6px; border: 1px solid #cbd5e1; font-size: 10pt; }}
        tr:nth-child(even) {{ background-color: #f1f5f9; }}
        
        @page {{
            margin: 0.5cm;
            size: A4;
        }}
    </style>
    <title>طلبية {order.order_number}</title>
</head>
<body>
    <div class="header-title">تفاصيل الطلبية</div>
    
    <div class="info-grid">
        <div class="info-item">
            <span class="label">اسم المستلم:</span>
            <span class="value">{order.recipient_name or '-'}</span>
        </div>
        <div class="info-item">
            <span class="label">رقم الطلبية:</span>
            <span class="value" style="font-family: monospace;">{order.order_number}</span>
        </div>
        <div class="info-item">
            <span class="label">عدد المنتجات:</span>
            <span class="value">{order.total_products} منتج</span>
        </div>
        <div class="info-item">
            <span class="label">إجمالي الكميات المسحوبة:</span>
            <span class="value">{order.total_quantities} حبة</span>
        </div>
        <div class="info-item" style="grid-column: span 2;">
            <span class="label">تاريخ وساعة السحب:</span>
            <span class="value">{order.created_at.strftime('%Y-%m-%d %H:%M')}</span>
        </div>
    </div>

    <div style="font-weight: bold; font-size: 14pt; margin-bottom: 10px; color: #1e293b;">المنتجات في هذه الطلبية:</div>
    
    <table>
        <thead>
            <tr>
                <th style="width: 10%;">#</th>
                <th style="width: 30%;">رقم المنتج</th>
                <th style="width: 20%;">الصورة</th>
                <th style="width: 20%;">السعر</th>
                <th style="width: 20%;">الكمية المسحوبة</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
</body>
</html>
'''
        with sync_playwright() as p:
            # التحقق من المسار الافتراضي لمتصفحات Playwright في لينكس (للإنتاج)
            import os
            
            # محاولة العثور على المتصفح المثبت (في حالة النشر)
            browser_args = {
                'headless': True,
                'args': ['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
            }
            
            try:
                browser = p.chromium.launch(**browser_args)
            except Exception as launch_error:
                # في حالة فشل التشغيل، نحاول تثبيت المتصفح تلقائياً (حل أخير)
                if "Executable doesn't exist" in str(launch_error):
                    import subprocess
                    subprocess.run(["playwright", "install", "chromium"], check=True)
                    browser = p.chromium.launch(**browser_args)
                else:
                    raise launch_error
                    
            page = browser.new_page()
            
            # معالجة الصور وتحويلها إلى base64 لضمان ظهورها
            # هذا يحل مشكلة عدم ظهور الصور في PDF في الإنتاج
            processed_html = html_content
            
            page.set_content(processed_html)
            pdf_bytes = page.pdf(
                format='A4',
                print_background=True,
                margin={'top': '1cm', 'right': '1cm', 'bottom': '1cm', 'left': '1cm'}
            )
            browser.close()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f"order_{order.order_number}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        import traceback
        error_msg = f'خطأ في إنشاء PDF: {str(e)}\n{traceback.format_exc()}'
        return HttpResponse(error_msg, content_type='text/plain')


def export_products_excel(request):
    """تصدير قائمة المنتجات إلى Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # إنشاء workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "المنتجات"
    
    # تنسيق النمط العربي
    arabic_font = Font(name='Arial', size=12, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # رأس الجدول
    headers = ['#', 'رقم المنتج', 'الاسم', 'الفئة', 'الكمية', 'الموقع', 'تاريخ الإضافة']
    ws.append(headers)
    
    # تنسيق رأس الجدول
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = border
    
    # جلب المنتجات
    products = Product.objects.all().order_by('product_number')
    
    # إضافة البيانات
    for idx, product in enumerate(products, start=1):
        location_text = product.location.full_location if product.location else 'لا يوجد موقع'
        
        row_data = [
            idx,
            product.product_number,
            product.name,
            product.category or '',
            product.quantity,
            location_text,
            product.created_at.strftime('%Y-%m-%d')
        ]
        
        ws.append(row_data)
        
        # تنسيق الصف
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=idx + 1, column=col)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = Font(name='Arial', size=11)
    
    # ضبط عرض الأعمدة
    column_widths = [5, 15, 25, 15, 15, 10, 12, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="قائمة_المنتجات.xlsx"'
    
    wb.save(response)
    return response


 


 


def convert_to_hijri(gregorian_date):
    """تحويل التاريخ الميلادي إلى هجري باستخدام حساب أدق"""
    try:
        from datetime import datetime
        import math
        
        # التاريخ المرجعي: 16 يوليو 622 ميلادي = 1 محرم 1 هجري
        gregorian_start = datetime(622, 7, 16)
        hijri_start = 1  # سنة 1 هجري
        
        # حساب الفرق بالأيام
        date_obj = datetime(gregorian_date.year, gregorian_date.month, gregorian_date.day)
        days_diff = (date_obj - gregorian_start).days
        
        # حساب السنة الهجرية (السنة الهجرية = 354.37 يوم في المتوسط)
        # مع تعديل أدق
        hijri_year = 1 + int(days_diff / 354.367)
        
        # حساب الأيام المتبقية منذ بداية السنة
        days_from_start_of_year = days_diff % 354
        
        # أشهر السنة الهجرية مع عدد أيامها (تقريبي)
        hijri_months = [
            ('محرم', 30), ('صفر', 29), ('ربيع الأول', 30), ('ربيع الآخر', 29),
            ('جمادى الأولى', 30), ('جمادى الآخرة', 29), ('رجب', 30), ('شعبان', 29),
            ('رمضان', 30), ('شوال', 29), ('ذو القعدة', 30), ('ذو الحجة', 29)
        ]
        
        # حساب الشهر واليوم
        remaining_days = days_from_start_of_year
        hijri_month = 1
        hijri_day = 1
        
        for month_name, month_days in hijri_months:
            if remaining_days < month_days:
                hijri_day = remaining_days + 1
                break
            remaining_days -= month_days
            hijri_month += 1
        
        # ضمان أن الشهر ضمن النطاق الصحيح
        if hijri_month > 12:
            hijri_month = 12
            hijri_day = min(hijri_day, 29)
        
        month_name = hijri_months[hijri_month - 1][0]
        
        # حساب السنة الهجرية بدقة أكبر باستخدام سنة كبيسة
        # السنة الهجرية الكبيسة لها 355 يوماً (3 سنوات في كل 8)
        leap_days = int(hijri_year / 30) * 11  # كل 30 سنة = 11 يوم إضافي
        adjusted_days = days_diff - leap_days
        
        # إعادة حساب السنة
        final_year = int(adjusted_days / 354.367) + 1
        
        return f"{hijri_day} {month_name} {final_year} هـ"
    except Exception as e:
        # في حالة الخطأ، رجع التاريخ الميلادي
        return f"{gregorian_date.strftime('%Y-%m-%d')}"


 


 


 

 


@exclude_maintenance
@login_required
def backup_restore_page(request):
    """صفحة النسخ الاحتياطي والاستعادة - للمسؤول فقط"""
    # إحصائيات البيانات الشاملة
    stats = {
        'warehouses': Warehouse.objects.count(),
        'locations': Location.objects.count(),
        'products': Product.objects.count(),
        'audit_logs': AuditLog.objects.count(),
        'orders': Order.objects.count(),
        'returns': ProductReturn.objects.count(),
        'user_profiles': UserProfile.objects.count(),
        'user_activity_logs': UserActivityLog.objects.count(),
    }
    
    return render(request, 'inventory_app/backup_restore.html', {
        'stats': stats
    })


@csrf_exempt
@require_http_methods(["POST"])
def update_location_notes(request):
    """تحديث ملاحظات موقع"""
    try:
        data = json.loads(request.body)
        location_id = data.get('location_id')
        notes = data.get('notes', '')
        
        if not location_id:
            return JsonResponse({
                'success': False,
                'error': 'معرف الموقع مطلوب'
            })
        
        location = Location.objects.get(id=location_id)
        location.notes = notes
        location.save()
        
        return JsonResponse({
            'success': True,
            'message': 'تم حفظ الملاحظات بنجاح'
        })
        
    except Location.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'الموقع غير موجود'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_http_methods(["POST"])
@exclude_maintenance
@login_required
def inspect_backup(request):
    """تحليل ملف النسخ الاحتياطي وإرجاع تقرير تفصيلي قبل الاستيراد"""
    try:
        uploaded_file = request.FILES.get('backup_file')
        inline_json = request.POST.get('backup_json')
        if not uploaded_file and not inline_json:
            return JsonResponse({'success': False, 'error': 'لم يتم إرسال ملف أو محتوى JSON'})

        filename = getattr(uploaded_file, 'name', '') if uploaded_file else 'inline.json'
        size = getattr(uploaded_file, 'size', None)
        raw_bytes = uploaded_file.read() if uploaded_file else inline_json.encode('utf-8', errors='ignore')
        data, parse_meta = _load_backup_data(raw_bytes, filename)
        if data is None:
            msg = 'الملف غير صالح (JSON)'
            if isinstance(parse_meta, dict) and parse_meta.get('message'):
                msg = msg + ' - ' + parse_meta.get('message')
            return JsonResponse({'success': False, 'error': msg})
        if isinstance(data, list):
            grouped = {}
            for item in data:
                if isinstance(item, dict):
                    model = item.get('model') or ''
                    name = model.split('.')[-1] if model else ''
                    if not name:
                        continue
                    section = name + 's'
                    if name == 'userprofile':
                        section = 'user_profiles'
                    elif name == 'useractivitylog':
                        section = 'user_activity_logs'
                    elif name == 'auditlog':
                        section = 'audit_logs'
                    grouped.setdefault(section, []).append(item)
            data = {'export_info': {'description': 'array_payload_transformed'}, **grouped}

        export_info = data.get('export_info', {})
        known_sections = [
            'warehouses', 'locations', 'products',
            'orders', 'returns', 'audit_logs',
            'user_profiles', 'user_activity_logs'
        ]
        sections = []
        schema_errors = []

        def analyze_section(name, items):
            info = {
                'name': name,
                'count': 0,
                'models': {},
                'schema_ok': True,
                'errors': [],
                'sample': []
            }
            if not isinstance(items, list):
                info['schema_ok'] = False
                info['errors'].append('القسم ليس مصفوفة')
                return info
            info['count'] = len(items)
            for idx, item in enumerate(items):
                if idx < 3:
                    try:
                        info['sample'].append(item)
                    except Exception:
                        pass
                if not isinstance(item, dict):
                    info['schema_ok'] = False
                    info['errors'].append(f'{name}[{idx}]: العنصر ليس كائناً JSON')
                    continue
                for req in ['model', 'pk', 'fields']:
                    if req not in item:
                        info['schema_ok'] = False
                        info['errors'].append(f"{name}[{idx}]: المكوّن '{req}' مفقود")
                        break
                model = item.get('model')
                if isinstance(model, str):
                    info['models'][model] = info['models'].get(model, 0) + 1
            return info

        for key, value in data.items():
            if isinstance(value, list):
                sec = analyze_section(key, value)
                sections.append(sec)
                if sec['errors']:
                    schema_errors.extend(sec['errors'])

        # تحليل التكرار في المنتجات الموجودة مسبقاً
        product_duplicates = []
        if 'products' in data and isinstance(data['products'], list):
            try:
                from collections import Counter
                backup_numbers = []
                for item in data['products']:
                    if isinstance(item, dict) and isinstance(item.get('fields'), dict):
                        pn = item['fields'].get('product_number')
                        if pn:
                            backup_numbers.append(pn)
                counts = Counter(backup_numbers)
                numbers = list(counts.keys())
                existing = Product.objects.filter(product_number__in=numbers).values('product_number', 'name')
                exist_map = {e['product_number']: e['name'] for e in existing}
                for pn, cnt in counts.items():
                    if pn in exist_map or cnt > 1:
                        product_duplicates.append({
                            'product_number': pn,
                            'count_in_backup': cnt,
                            'existing': pn in exist_map,
                            'existing_name': exist_map.get(pn)
                        })
            except Exception:
                pass

        # اقتراح التبعيات
        present = set([s['name'] for s in sections if s['count'] > 0])
        suggested_dependencies = {}
        if 'products' in present:
            suggested_dependencies['products'] = ['locations', 'warehouses']
        if 'locations' in present:
            suggested_dependencies['locations'] = ['warehouses']
        if 'audit_logs' in present:
            suggested_dependencies['audit_logs'] = ['products', 'locations', 'warehouses']

        return JsonResponse({
            'success': True,
            'filename': filename,
            'size': size,
            'export_info': export_info,
            'sections': sections,
            'suggested_dependencies': suggested_dependencies,
            'has_errors': len(schema_errors) > 0,
            'schema_errors': schema_errors[:100],
            'product_duplicates': product_duplicates,
            'parse_meta': parse_meta
        }, json_dumps_params={'ensure_ascii': False})

    except json.JSONDecodeError as e:
        return JsonResponse({'success': False, 'error': f'الملف غير صالح (JSON): {str(e)}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
@exclude_maintenance
@login_required
def reset_environment(request):
    """حذف كل البيانات لتجهيز بيئة نظيفة (يستثني المشرفين)"""
    try:
        # حذف بالترتيب للتأكد من العلاقات
        UserActivityLog.objects.all().delete()
        AuditLog.objects.all().delete()
        ProductReturn.objects.all().delete()
        Order.objects.all().delete()
        Product.objects.all().delete()
        Location.objects.all().delete()
        Warehouse.objects.all().delete()
        UserProfile.objects.exclude(user__is_superuser=True).delete()

        return JsonResponse({'success': True, 'message': 'تم تهيئة بيئة نظيفة بنجاح'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
def export_backup(request):
    """تصدير النسخ الاحتياطي الشامل - يشمل جميع البيانات والأنشطة"""
    try:
        # جمع جميع البيانات بشكل شامل
        data = {
            'export_info': {
                'date': datetime.now().isoformat(),
                'version': '2.0',
                'description': 'نسخ احتياطي شامل كامل من نظام إدارة المستودع - يشمل جميع البيانات والعمليات والأنشطة'
            },
            # البيانات الأساسية
            'warehouses': json.loads(serializers.serialize('json', Warehouse.objects.all())),
            'locations': json.loads(serializers.serialize('json', Location.objects.all())),
            'products': json.loads(serializers.serialize('json', Product.objects.all())),
            
            # سجلات العمليات
            'audit_logs': json.loads(serializers.serialize('json', AuditLog.objects.all())),
            'orders': json.loads(serializers.serialize('json', Order.objects.all())),
            'returns': json.loads(serializers.serialize('json', ProductReturn.objects.all())),
            
        # التقارير (تمت إزالة التقارير اليومية)
            
            # بيانات المستخدمين والأنشطة (استثناء admin من UserProfile لتجنب التكرار)
            # نحذف UserProfile للمستخدمين الذين هم superuser (admin) لتجنب التكرار عند الاستيراد
            'user_profiles': json.loads(serializers.serialize('json', UserProfile.objects.exclude(user__is_superuser=True))),
            'user_activity_logs': json.loads(serializers.serialize('json', UserActivityLog.objects.all())),
            
            # إحصائيات النسخ الاحتياطي
            'backup_stats': {
                'warehouses_count': Warehouse.objects.count(),
                'locations_count': Location.objects.count(),
                'products_count': Product.objects.count(),
                'audit_logs_count': AuditLog.objects.count(),
                'orders_count': Order.objects.count(),
                'returns_count': ProductReturn.objects.count(),
                
                'user_profiles_count': UserProfile.objects.exclude(user__is_superuser=True).count(),  # عدد UserProfiles بدون admin
                'user_activity_logs_count': UserActivityLog.objects.count(),
            }
        }
        
        # إنشاء ملف JSON
        json_data = json.dumps(data, ensure_ascii=False, indent=2)
        
        # إعداد الاستجابة
        filename = f'backup_full_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f'Error in export_backup: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_http_methods(["POST"])
@exclude_maintenance
@login_required
def import_backup(request):
    """استيراد النسخ الاحتياطي"""
    try:
        uploaded_file = request.FILES.get('backup_file')
        inline_json = request.POST.get('backup_json')
        if not uploaded_file and not inline_json:
            return JsonResponse({'success': False, 'error': 'لم يتم إرسال ملف أو محتوى JSON'})
        
        filename = getattr(uploaded_file, 'name', '') if uploaded_file else 'inline.json'
        raw_bytes = uploaded_file.read() if uploaded_file else inline_json.encode('utf-8', errors='ignore')
        data, parse_meta = _load_backup_data(raw_bytes, filename)
        if data is None:
            msg = 'الملف غير صالح (JSON)'
            if isinstance(parse_meta, dict) and parse_meta.get('message'):
                msg = msg + ' - ' + parse_meta.get('message')
            return JsonResponse({'success': False, 'error': msg})
        if isinstance(data, list):
            grouped = {}
            for item in data:
                if isinstance(item, dict):
                    model = item.get('model') or ''
                    name = model.split('.')[-1] if model else ''
                    if not name:
                        continue
                    section = name + 's'
                    if name == 'userprofile':
                        section = 'user_profiles'
                    elif name == 'useractivitylog':
                        section = 'user_activity_logs'
                    elif name == 'auditlog':
                        section = 'audit_logs'
                    grouped.setdefault(section, []).append(item)
            data = {'export_info': {'description': 'array_payload_transformed'}, **grouped}
        
        # تهيئة export_info إن كان مفقوداً
        if 'export_info' not in data or not isinstance(data.get('export_info'), dict):
            data['export_info'] = {
                'date': datetime.now().isoformat(),
                'version': 'unknown',
                'description': 'استيراد بدون معلومات التصدير'
            }
        
        clear_existing = request.POST.get('clear_existing', 'false') == 'true'
        avoid_duplicates = request.POST.get('avoid_duplicates', 'false') == 'true'
        # الأقسام المحددة (اختياري)
        selected_sections_raw = request.POST.get('selected_sections', '')
        try:
            selected_sections = set(json.loads(selected_sections_raw)) if selected_sections_raw else set()
        except Exception:
            selected_sections = set()
        # إذا لم تُحدّد أقسام، نستخدم الافتراضي: كل ما هو موجود في الملف
        if not selected_sections:
            selected_sections = set([k for k, v in data.items() if isinstance(v, list)])
        
        # تحقق بنية الأقسام قبل الاستيراد وإعداد إحصائيات
        schema_errors = []
        for section in selected_sections:
            if section in data:
                if not isinstance(data[section], list):
                    schema_errors.append(f"القسم '{section}' يجب أن يكون مصفوفة")
                    continue
                for idx, item in enumerate(data[section]):
                    if not isinstance(item, dict):
                        schema_errors.append(f"{section}[{idx}]: العنصر ليس كائناً JSON")
                        break
                    for k in ['model', 'pk', 'fields']:
                        if k not in item:
                            schema_errors.append(f"{section}[{idx}]: المكوّن '{k}' مفقود")
                            break
        if schema_errors:
            return JsonResponse({'success': False, 'error': 'أخطاء في بنية الملف', 'details': schema_errors})
        
        # إضافة التبعيات اللازمة تلقائياً لتجنب أخطاء المراجع
        # المنتجات تحتاج الأماكن والمستودعات، والأماكن تحتاج مستودعات، والسجلات تحتاج المنتجات
        if 'products' in selected_sections:
            selected_sections.update(['locations', 'warehouses'])
        if 'locations' in selected_sections:
            selected_sections.add('warehouses')
        if 'audit_logs' in selected_sections:
            selected_sections.update(['products', 'locations', 'warehouses'])
        
        # بدء الاستيراد
        import_counts = {s: 0 for s in selected_sections}
        import_errors = []
        with transaction.atomic():
            if clear_existing:
                # حذف بحسب الأقسام المختارة، مع مراعاة العلاقات
                if 'user_activity_logs' in selected_sections:
                    UserActivityLog.objects.all().delete()
                if 'audit_logs' in selected_sections:
                    AuditLog.objects.all().delete()
                if 'returns' in selected_sections:
                    ProductReturn.objects.all().delete()
                if 'orders' in selected_sections:
                    Order.objects.all().delete()
                if 'products' in selected_sections:
                    Product.objects.all().delete()
                if 'locations' in selected_sections:
                    Location.objects.all().delete()
                if 'warehouses' in selected_sections:
                    Warehouse.objects.all().delete()
                if 'user_profiles' in selected_sections:
                    # حماية الحساب المسؤول 'ammar' من الحذف أثناء الاستيراد
                    UserProfile.objects.exclude(user__username='ammar').exclude(user__is_superuser=True).delete()
            
            # استيراد البيانات بالترتيب الصحيح (حسب العلاقات)
            # 1. البيانات الأساسية أولاً
            if 'warehouses' in selected_sections and 'warehouses' in data:
                objects = serializers.deserialize('json', json.dumps(data['warehouses']))
                for i, obj in enumerate(objects):
                    try:
                        obj.save()
                        import_counts['warehouses'] += 1
                    except Exception as e:
                        import_errors.append(f"warehouses[{i}]: {str(e)}")
            
            if 'locations' in selected_sections and 'locations' in data:
                objects = serializers.deserialize('json', json.dumps(data['locations']))
                for i, obj in enumerate(objects):
                    try:
                        obj.save()
                        import_counts['locations'] += 1
                    except Exception as e:
                        import_errors.append(f"locations[{i}]: {str(e)}")
            
            if 'products' in selected_sections and 'products' in data:
                objects = serializers.deserialize('json', json.dumps(data['products']))
                for i, obj in enumerate(objects):
                    try:
                        instance = obj.object
                        if avoid_duplicates:
                            pn = getattr(instance, 'product_number', None)
                            if pn:
                                existing = Product.objects.filter(product_number=pn).first()
                                if existing:
                                    instance.pk = existing.pk
                        instance.save()
                        import_counts['products'] += 1
                    except Exception as e:
                        import_errors.append(f"products[{i}]: {str(e)}")
            
            # 2. ملفات المستخدمين (معالجة ذكية للمستخدمين المفقودين)
            if 'user_profiles' in selected_sections and 'user_profiles' in data:
                # التأكد من وجود مستخدم النظام الافتراضي
                system_user, _ = User.objects.get_or_create(username='system', defaults={'email': 'system@local', 'is_active': False})
                
                for i, profile_data in enumerate(data['user_profiles']):
                    try:
                        # استخراج user_id من البيانات
                        user_id = None
                        if 'fields' in profile_data:
                            user_id = profile_data['fields'].get('user') or profile_data['fields'].get('user_id')
                        
                        if not user_id:
                            continue

                        # التحقق من وجود المستخدم
                        if not User.objects.filter(pk=user_id).exists():
                            # إذا كان المستخدم غير موجود، نربط الملف بمستخدم النظام أو نتجاوزه
                            # هنا سنقوم بإنشاء مستخدم وهمي للحفاظ على تكامل البيانات
                            try:
                                # محاولة استرداد اسم المستخدم القديم إذا كان متاحاً في مكان ما، أو استخدام ID
                                dummy_username = f"imported_user_{user_id}"
                                User.objects.create(pk=user_id, username=dummy_username, is_active=False)
                            except Exception:
                                # إذا فشل إنشاء المستخدم بنفس الـ ID (ربما تعارض)، نستخدم مستخدم النظام
                                profile_data['fields']['user'] = system_user.id
                        
                        # الآن نحاول الاستيراد
                        objects = serializers.deserialize('json', json.dumps([profile_data]))
                        for obj in objects:
                            # حماية: لا تقم بتحديث ملف المسؤول ammar إذا كان موجوداً في النسخة الاحتياطية بمعلومات قديمة
                            if obj.object.user.username == 'ammar':
                                continue
                            obj.save()
                            import_counts['user_profiles'] += 1
                            
                    except Exception as e:
                        import_errors.append(f"user_profiles[{i}]: {str(e)}")

            
            # 3. سجلات العمليات
            if 'audit_logs' in selected_sections and 'audit_logs' in data:
                objects = serializers.deserialize('json', json.dumps(data['audit_logs']))
                for i, obj in enumerate(objects):
                    try:
                        obj.save()
                        import_counts['audit_logs'] += 1
                    except Exception as e:
                        import_errors.append(f"audit_logs[{i}]: {str(e)}")
            
            if 'orders' in selected_sections and 'orders' in data:
                objects = serializers.deserialize('json', json.dumps(data['orders']))
                for i, obj in enumerate(objects):
                    try:
                        obj.save()
                        import_counts['orders'] += 1
                    except Exception as e:
                        import_errors.append(f"orders[{i}]: {str(e)}")
            
            if 'returns' in selected_sections and 'returns' in data:
                objects = serializers.deserialize('json', json.dumps(data['returns']))
                for i, obj in enumerate(objects):
                    try:
                        obj.save()
                        import_counts['returns'] += 1
                    except Exception as e:
                        import_errors.append(f"returns[{i}]: {str(e)}")
            
            # 4. التقارير (لا يوجد تقارير يومية بعد الإزالة)
            
            # 5. سجلات أنشطة المستخدمين
            if 'user_activity_logs' in selected_sections and 'user_activity_logs' in data:
                # التأكد من وجود مستخدم النظام لمعالجة السجلات التي ليس لها مستخدم
                system_user, _ = User.objects.get_or_create(username='system', defaults={'email': 'system@local', 'is_active': False})
                
                # معالجة السجلات قبل الاستيراد
                logs_data = data['user_activity_logs']
                for log in logs_data:
                    if 'fields' in log and 'user' in log['fields']:
                        user_id = log['fields']['user']
                        # التحقق من وجود المستخدم
                        if not User.objects.filter(pk=user_id).exists():
                            # ربط السجل بمستخدم النظام بدلاً من المستخدم المفقود
                            log['fields']['user'] = system_user.id

                objects = serializers.deserialize('json', json.dumps(logs_data))
                for i, obj in enumerate(objects):
                    try:
                        obj.save()
                        import_counts['user_activity_logs'] += 1
                    except Exception as e:
                        import_errors.append(f"user_activity_logs[{i}]: {str(e)}")
        
        if import_errors:
            return JsonResponse({
                'success': False,
                'error': 'حدثت أخطاء أثناء الاستيراد',
                'details': import_errors[:50],
                'imported': import_counts
            })
        return JsonResponse({
            'success': True,
            'message': 'تم الاستيراد بنجاح',
            'imported': import_counts
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'error': f'الملف غير صالح (JSON): {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@exclude_maintenance
@login_required
def data_deletion_page(request):
    """صفحة حذف البيانات"""
    # إحصائيات البيانات الشاملة
    stats = {
        'warehouses': Warehouse.objects.count(),
        'locations': Location.objects.count(),
        'products': Product.objects.count(),
        'audit_logs': AuditLog.objects.count(),
        'orders': Order.objects.count(),
        'returns': ProductReturn.objects.count(),
        'user_profiles': UserProfile.objects.count(),
        'user_activity_logs': UserActivityLog.objects.count(),
    }
    
    return render(request, 'inventory_app/data_deletion.html', {
        'stats': stats
    })


@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
@exclude_maintenance
@login_required
def delete_data(request):
    """حذف البيانات المحددة"""
    try:
        data = json.loads(request.body)
        # التحقق من كلمة المرور
        password = str(data.get('password', '') or '').strip()
        # ملاحظة أمنية: يفضل وضع كلمة المرور في الإعدادات/المتغيرات البيئية
        REQUIRED_PASSWORD = 'Thepest**1'
        if not password:
            return JsonResponse({'success': False, 'error': 'يجب إدخال كلمة مرور الحذف'}, status=400)
        if password != REQUIRED_PASSWORD:
            return JsonResponse({'success': False, 'error': 'كلمة مرور غير صحيحة'}, status=403)
        
        # قراءة البيانات المحددة للحذف
        delete_products = data.get('delete_products', False)
        delete_locations = data.get('delete_locations', False)
        delete_warehouses = data.get('delete_warehouses', False)
        delete_audit_logs = data.get('delete_audit_logs', False)
        
        delete_orders = data.get('delete_orders', False)
        delete_returns = data.get('delete_returns', False)
        delete_user_profiles = data.get('delete_user_profiles', False)
        delete_user_activity_logs = data.get('delete_user_activity_logs', False)
        
        deleted_items = []
        
        # حذف البيانات المحددة بالترتيب الصحيح (تجنب أخطاء Foreign Key)
        # 1. حذف السجلات التي تعتمد على بيانات أخرى أولاً
        if delete_user_activity_logs:
            count = UserActivityLog.objects.count()
            UserActivityLog.objects.all().delete()
            deleted_items.append(f'{count} سجل نشاط مستخدم')
        
        if delete_audit_logs:
            count = AuditLog.objects.count()
            AuditLog.objects.all().delete()
            deleted_items.append(f'{count} سجل عمليات')
        
        if delete_returns:
            count = ProductReturn.objects.count()
            ProductReturn.objects.all().delete()
            deleted_items.append(f'{count} مرتجع')
        
        if delete_orders:
            count = Order.objects.count()
            Order.objects.all().delete()
            deleted_items.append(f'{count} طلبية')
        
        if delete_products:
            count = Product.objects.count()
            Product.objects.all().delete()
            deleted_items.append(f'{count} منتج')
        
        if delete_locations:
            count = Location.objects.count()
            Location.objects.all().delete()
            deleted_items.append(f'{count} مكان')
        
        if delete_warehouses:
            count = Warehouse.objects.count()
            Warehouse.objects.all().delete()
            deleted_items.append(f'{count} مستودع')
        
        
        
        if delete_user_profiles:
            # حذف جميع ملفات المستخدمين ما عدا المستخدم المسؤول 'ammar'
            profiles_to_delete = UserProfile.objects.exclude(user__username='ammar')
            count = profiles_to_delete.count()
            profiles_to_delete.delete()
            
            # حذف المستخدمين المرتبطين (ما عدا ammar)
            users_to_delete = User.objects.exclude(username='ammar').exclude(is_superuser=True)
            users_count = users_to_delete.count()
            users_to_delete.delete()
            
            deleted_items.append(f'{count} ملف مستخدم و {users_count} حساب')
        
        if not deleted_items:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم تحديد أي بيانات للحذف'
            })
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف: {", ".join(deleted_items)}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'البيانات غير صالحة'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def get_all_recipients_stats(request):
    """API لجلب إحصائيات جميع المستلمين"""
    from django.db.models import Count, Sum
    recipients = (
        Order.objects.exclude(recipient_name__isnull=True)
        .exclude(recipient_name='')
        .values('recipient_name')
        .annotate(count=Count('id'), total_qty=Sum('total_quantities'))
        .order_by('-count')
    )
    return JsonResponse({
        'success': True,
        'recipients': list(recipients)
    })


@login_required
def orders_list(request):
    """عرض قائمة الطلبات المسحوبة مع إحصائيات متقدمة"""
    from django.core.paginator import Paginator
    from django.db.models.functions import TruncDate
    from django.db.models import Count, Sum
    import json
    
    # الاستعلام الأساسي
    orders_qs = Order.objects.all().order_by('-created_at')
    
    # تصفية حسب المستلم إذا تم تحديده
    recipient_filter = request.GET.get('recipient')
    product_query = request.GET.get('product_query', '').strip().lower()
    recipient_items = []
    
    if recipient_filter:
        orders_qs = orders_qs.filter(recipient_name=recipient_filter)
        
        # تجميع سجل المواد للمستلم (لعرض جدول تفصيلي)
        # نستخدم all() للحصول على كل العناصر، يمكن تحديد العدد لاحقاً إذا كان الأداء بطيئاً
        for order in orders_qs:
            products = order.products_data
            if isinstance(products, str):
                try:
                    import json
                    products = json.loads(products)
                except:
                    products = []
            
            for item in products:
                # إذا كان هناك بحث محدد عن منتج، نتجاهل المنتجات الأخرى
                if product_query:
                    p_num = str(item.get('product_number', '')).strip().lower()
                    # بحث دقيق لرقم المنتج
                    if product_query != p_num:
                        continue

                qty = item.get('quantity_taken')
                if qty is None:
                    qty = item.get('quantity', 0)
                
                if int(qty) > 0:
                    recipient_items.append({
                        'date': order.created_at,
                        'order_number': order.order_number,
                        'order_id': order.id,
                        'product_name': item.get('name', 'منتج'),
                        'product_number': item.get('product_number', ''),
                        'quantity': int(qty),
                    })
        
        # ترتيب حسب التاريخ الأحدث
        recipient_items.sort(key=lambda x: x['date'], reverse=True)
    
    # إحصائيات عامة
    total_orders = Order.objects.count()
    today_orders = Order.objects.filter(created_at__date=timezone.now().date()).count()
    total_quantities_taken = Order.objects.aggregate(
        total=db_models.Sum('total_quantities')
    )['total'] or 0

    # 1. إحصائيات الرسم البياني: حركة السحب اليومية (آخر 30 يوم)
    last_30_days = timezone.now() - timedelta(days=30)
    daily_stats = (
        Order.objects.filter(created_at__gte=last_30_days)
        .annotate(date=TruncDate('created_at'))
        .values('date')
        .annotate(count=Count('id'), qty=Sum('total_quantities'))
        .order_by('date')
    )
    
    daily_labels = [s['date'].strftime('%Y-%m-%d') for s in daily_stats]
    daily_counts = [s['count'] for s in daily_stats]
    daily_qtys = [s['qty'] for s in daily_stats]

    # 2. إحصائيات الرسم البياني: أكثر المستلمين نشاطاً (Top 5)
    top_recipients = (
        Order.objects.exclude(recipient_name__isnull=True)
        .exclude(recipient_name='')
        .values('recipient_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    
    recipient_labels = [s['recipient_name'] for s in top_recipients]
    recipient_counts = [s['count'] for s in top_recipients]
    
    # إضافة Pagination - 20 طلب لكل صفحة
    paginator = Paginator(orders_qs, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'orders': page_obj,
        'page_obj': page_obj,
        'total_orders': total_orders,
        'today_orders': today_orders,
        'total_quantities_taken': total_quantities_taken,
        # بيانات الرسوم البيانية
        'daily_labels': json.dumps(daily_labels),
        'daily_counts': json.dumps(daily_counts),
        'daily_qtys': json.dumps(daily_qtys),
        'recipient_labels': json.dumps(recipient_labels),
        'recipient_counts': json.dumps(recipient_counts),
        'recipient_items': recipient_items, # السجل التفصيلي للمستلم
    }
    
    return render(request, 'inventory_app/orders_list.html', context)


@login_required
def search_order_history(request):
    """API للبحث الذكي في سجلات المنتجات داخل الطلبات"""
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'success': False, 'error': 'No query provided'})
    
    # 1. البحث الأولي السريع: الطلبات التي يحتوي نصها (JSON) على كلمة البحث
    candidates = Order.objects.filter(products_data__icontains=query)
    
    total_withdrawn = 0
    recipients_map = {} # name -> count
    dates_map = {} # date -> quantity
    
    # 2. الفحص الدقيق وحساب الكميات
    for order in candidates:
        try:
            # products_data هو قائمة من القواميس
            products = order.products_data
            if isinstance(products, str):
                import json
                products = json.loads(products)
                
            for p in products:
                # التحقق مما إذا كان هذا المنتج هو المقصود
                p_name = str(p.get('name', '')).lower()
                p_num = str(p.get('product_number', '')).strip().lower()
                q_lower = query.lower()
                
                # بحث دقيق لرقم المنتج، أو جزئي للاسم
                if q_lower == p_num or q_lower in p_name:
                    # قد يكون المفتاح quantity أو quantity_taken حسب إصدار البيانات
                    qty = int(p.get('quantity_taken', p.get('quantity', 0)))
                    total_withdrawn += qty
                    
                    # تجميع المستلمين
                    r_name = order.recipient_name or 'غير محدد'
                    # نقوم بجمع الكميات بدلاً من عدد الطلبات
                    recipients_map[r_name] = recipients_map.get(r_name, 0) + qty
                    
                    # تجميع التواريخ
                    date_str = order.created_at.strftime('%Y-%m-%d')
                    dates_map[date_str] = dates_map.get(date_str, 0) + qty
        except Exception:
            continue
            
    # تحضير النتائج
    # إرجاع جميع المستلمين (بدون تقييد بـ 5)
    sorted_recipients = sorted(recipients_map.items(), key=lambda x: x[1], reverse=True)
    sorted_dates = sorted(dates_map.items())
    
    return JsonResponse({
        'success': True,
        'stats': {
            'total_withdrawn': total_withdrawn,
            'top_recipients': [{'name': k, 'count': v} for k, v in sorted_recipients],
            'timeline': [{'date': k, 'qty': v} for k, v in sorted_dates]
        }
    })


@login_required
def order_detail(request, order_id):
    """عرض تفاصيل طلبية محددة"""
    order = get_object_or_404(Order, id=order_id)
    
    # تصفية المنتجات إذا كان هناك بحث
    product_query = request.GET.get('product_query', '').strip().lower()
    
    # تأكد من أن products_data هو قائمة (إذا كان نصاً)
    if isinstance(order.products_data, str):
        try:
            import json
            order.products_data = json.loads(order.products_data)
        except:
            order.products_data = []

    # إذا كان هناك بحث، نقوم بتصفية القائمة
    if product_query:
        filtered_products = []
        for p in order.products_data:
            p_num = str(p.get('product_number', '')).strip().lower()
            # بحث دقيق
            if p_num == product_query:
                filtered_products.append(p)
        order.products_data = filtered_products

    # إثراء البيانات بالصور والسعر من قاعدة البيانات
    product_numbers = [p.get('product_number') for p in order.products_data if p.get('product_number')]
    products = Product.objects.filter(product_number__in=product_numbers)
    products_map = {p.product_number: p for p in products}
    
    for p in order.products_data:
        p_num = p.get('product_number')
        if p_num in products_map:
            product = products_map[p_num]
            p['price'] = float(product.price) if product.price is not None else 0
            if product.image:
                p['image_url'] = product.image.url
            elif product.image_url:
                p['image_url'] = product.image_url
            else:
                p['image_url'] = None
        else:
             p['price'] = 0

    return render(request, 'inventory_app/order_detail.html', {
        'order': order
    })


@csrf_exempt
def delete_order(request, order_id):
    """حذف طلبية محددة"""
    if request.method == 'DELETE':
        try:
            order = get_object_or_404(Order, id=order_id)
            order_number = order.order_number
            order.delete()
            return JsonResponse({
                'success': True,
                'message': f'تم حذف الطلبية {order_number}'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    return JsonResponse({'error': 'Invalid request method'}, status=400)





@login_required
@require_http_methods(["POST"])
@csrf_exempt
@transaction.atomic
def reset_all_quantities(request):
    try:
        import json
        from django.conf import settings
        body = {}
        if request.body:
            try:
                body = json.loads(request.body)
            except Exception:
                body = {}
        pwd = body.get('password')
        if not settings.RESET_PASSWORD or pwd != settings.RESET_PASSWORD:
            return JsonResponse({'success': False, 'error': 'كلمة المرور غير صحيحة'}, status=403)
        # جلب جميع المنتجات
        total_products = Product.objects.count()
        
        if total_products == 0:
            return JsonResponse({
                'success': True,
                'message': 'لا توجد منتجات في قاعدة البيانات',
                'count': 0
            })
        
        # تحديث جميع الكميات إلى 0
        updated_count = Product.objects.update(quantity=0)
        
        return JsonResponse({
            'success': True,
            'message': f'تم تصفير الكميات لجميع المنتجات بنجاح ({updated_count} منتج)',
            'count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ أثناء تصفير الكميات: {str(e)}'
        }, status=500)





# ========== نظام المستخدمين والصلاحيات ==========

@never_cache
@require_http_methods(["GET", "POST"])
@ratelimit(key='ip', rate='5/m', method='POST', block=True) if RATELIMIT_AVAILABLE else lambda x: x
def custom_login(request):
    """تسجيل الدخول مخصص مع تسجيل النشاط وتحسينات الأمان"""
    if request.user.is_authenticated:
        return redirect('inventory_app:home')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            
            
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                # التحقق من أن المستخدم نشط
                if not user.is_active:
                    security_logger.warning(f'محاولة تسجيل دخول لحساب معطل: {username} من IP: {request.META.get("REMOTE_ADDR")}')
                    messages.error(request, 'هذا الحساب معطل. يرجى التواصل مع المسؤول')
                    return render(request, 'auth/login.html', {'form': form})
                
                auth_login(request, user)
                
                # تسجيل نشاط تسجيل الدخول
                UserActivityLog.log_activity(
                    user=user,
                    action='login',
                    description=f'تم تسجيل الدخول بنجاح',
                    request=request
                )
                
                logger.info(f'User {username} logged in successfully from IP: {request.META.get("REMOTE_ADDR")}')
                
                # إنشاء UserProfile إذا لم يكن موجوداً
                if not hasattr(user, 'user_profile'):
                    UserProfile.objects.create(
                        user=user,
                        user_type='admin' if user.is_superuser else 'staff'
                    )
                
                messages.success(request, f'مرحباً {user.username}!')
                return redirect('inventory_app:home')
            else:
                # تسجيل محاولة فاشلة للأمان
                security_logger.warning(f'Failed login attempt for username: {username} from IP: {request.META.get("REMOTE_ADDR")}')
                messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')
                
                # تسجيل محاولة فاشلة
                try:
                    failed_user = User.objects.get(username=username)
                    UserActivityLog.log_activity(
                        user=failed_user,
                        action='login',
                        description=f'محاولة تسجيل دخول فاشلة - كلمة مرور خاطئة',
                        request=request
                    )
                except User.DoesNotExist:
                    pass
                
                return render(request, 'auth/login.html', {'form': form})
        else:
            # تسجيل أخطاء التحقق
            security_logger.warning(f'Invalid form data in login attempt from IP: {request.META.get("REMOTE_ADDR")}')
            messages.error(request, 'يرجى التحقق من البيانات المدخلة')
    
    else:
        form = LoginForm()
    
    return render(request, 'auth/login.html', {'form': form})


@login_required
def custom_logout(request):
    """تسجيل الخروج مخصص مع تسجيل النشاط"""
    user = request.user
    
    # تسجيل نشاط تسجيل الخروج
    UserActivityLog.log_activity(
        user=user,
        action='logout',
        description='تم تسجيل الخروج بنجاح',
        request=request
    )
    
    auth_logout(request)
    messages.success(request, 'تم تسجيل الخروج بنجاح')
    return redirect('login')


 

 


 

 

 

 

 

 

 

 


@admin_required
@never_cache
@require_http_methods(["GET", "POST"])
@ratelimit(key='user', rate='100/h', method='POST', block=True) if RATELIMIT_AVAILABLE else lambda x: x
def register_staff(request):
    """إنشاء حساب موظف جديد - للمسؤول فقط مع تحسينات الأمان"""
    if request.method == 'POST':
        form = RegisterStaffForm(request.POST)
        
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            email = form.cleaned_data['email']
            phone = form.cleaned_data['phone']
            user_type = 'staff'  # تعيين تلقائي كـ موظف
            
            try:
                # إنشاء المستخدم
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        email=email if email else None,
                        is_staff=False  # لا نجعله staff في Django default
                    )
                
                    # إنشاء UserProfile
                    UserProfile.objects.create(
                        user=user,
                        user_type=user_type,
                        phone=phone,
                        is_active=True
                    )
                
                # تسجيل النشاط
                UserActivityLog.log_activity(
                    user=request.user,
                    action='user_created',
                    description=f'تم إنشاء حساب {user_type} جديد: {username}',
                    request=request,
                    object_type='User',
                    object_id=user.id,
                    object_name=username
                )
                
                logger.info(f'Admin {request.user.username} created new staff account: {username}')
                messages.success(request, f'تم إنشاء حساب {username} بنجاح')
                return redirect('inventory_app:admin_dashboard')
                
            except Exception as e:
                security_logger.error(f'Error creating staff account: {str(e)} from IP: {request.META.get("REMOTE_ADDR")}')
                logger.error(f'Error in register_staff: {str(e)}')
                messages.error(request, f'حدث خطأ أثناء إنشاء الحساب: {str(e)}')
                return render(request, 'auth/register.html', {'form': form})
        else:
            # تسجيل أخطاء التحقق
            security_logger.warning(f'Invalid form data in register_staff from IP: {request.META.get("REMOTE_ADDR")}')
            messages.error(request, 'يرجى التحقق من البيانات المدخلة:')
            for field, errors in form.errors.items():
                # Get field label if available
                field_label = field
                if field in form.fields:
                    field_label = form.fields[field].label or field
                
                for error in errors:
                    messages.error(request, f"- {field_label}: {error}")
    else:
        form = RegisterStaffForm()
    
    return render(request, 'auth/register.html', {'form': form})


@admin_required
@exclude_admin_dashboard
def admin_dashboard(request):
    """لوحة تحكم المسؤول - عرض تتبع الموظفين"""
    # إحصائيات عامة
    total_staff = UserProfile.objects.filter(user_type='staff', is_active=True).count()
    total_admins = UserProfile.objects.filter(user_type='admin', is_active=True).count()
    
    # إحصائيات الأنشطة (آخر 7 أيام)
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_activities = UserActivityLog.objects.filter(created_at__gte=seven_days_ago).count()
    today_activities = UserActivityLog.objects.filter(
        created_at__date=timezone.now().date()
    ).count()
    
    # قائمة الموظفين مع إحصائياتهم (جميع الموظفين - نشط وغير نشط)
    # تحسين: استخدام aggregation بدلاً من queries متعددة
    from django.db.models import Count, Q, Prefetch
    
    staff_members = []
    staff_profiles = UserProfile.objects.filter(user_type='staff').select_related('user')
    
    # جلب جميع الأنشطة والعمليات دفعة واحدة لتجنب N+1
    today = timezone.now().date()
    
    for profile in staff_profiles:
        user = profile.user
        
        # استخدام aggregation للحصول على الإحصائيات بكفاءة
        activity_stats = UserActivityLog.objects.filter(user=user).aggregate(
            today_count=Count('id', filter=Q(created_at__date=today)),
            week_count=Count('id', filter=Q(created_at__gte=seven_days_ago))
        )
        
        operation_stats_agg = AuditLog.objects.filter(user=user.username).aggregate(
            today_operations=Count('id', filter=Q(created_at__date=today)),
            week_operations=Count('id', filter=Q(created_at__gte=seven_days_ago))
        )
        
        # آخر نشاط وعملية
        last_activity = UserActivityLog.objects.filter(user=user).order_by('-created_at').first()
        last_operation = AuditLog.objects.filter(user=user.username).select_related('product').order_by('-created_at').first()
        
        # العمليات الأخيرة (آخر 10 عمليات)
        recent_operations = AuditLog.objects.filter(user=user.username).select_related('product').order_by('-created_at')[:10]
        
        # إحصائيات العمليات حسب النوع - استخدام values وannotate
        operation_counts = AuditLog.objects.filter(user=user.username).values('action').annotate(
            count=Count('id')
        )
        operation_stats = {
            dict(AuditLog.ACTION_CHOICES).get(item['action'], item['action']): item['count']
            for item in operation_counts
        }
        
        staff_members.append({
            'profile': profile,
            'user': user,
            'today_activities': activity_stats['today_count'] or 0,
            'week_activities': activity_stats['week_count'] or 0,
            'today_operations': operation_stats_agg['today_operations'] or 0,
            'week_operations': operation_stats_agg['week_operations'] or 0,
            'last_activity': last_activity,
            'last_operation': last_operation,
            'recent_operations': recent_operations,
            'operation_stats': operation_stats,
            'last_login_ip': profile.last_login_ip,
        })
    
    # ترتيب الموظفين حسب النشاط
    staff_members.sort(key=lambda x: x['today_activities'], reverse=True)
    
    # قائمة المسؤولين مع إحصائياتهم (جميع المسؤولين - نشط وغير نشط)
    # تحسين: استخدام aggregation بدلاً من queries متعددة
    admin_members = []
    admin_profiles = UserProfile.objects.filter(user_type='admin').select_related('user')
    
    for profile in admin_profiles:
        user = profile.user
        
        # استخدام aggregation للحصول على الإحصائيات بكفاءة
        activity_stats_agg = UserActivityLog.objects.filter(user=user).aggregate(
            today_count=Count('id', filter=Q(created_at__date=today)),
            week_count=Count('id', filter=Q(created_at__gte=seven_days_ago))
        )
        
        operation_stats_agg = AuditLog.objects.filter(user=user.username).aggregate(
            today_operations=Count('id', filter=Q(created_at__date=today)),
            week_operations=Count('id', filter=Q(created_at__gte=seven_days_ago))
        )
        
        # آخر نشاط وعملية
        last_activity = UserActivityLog.objects.filter(user=user).order_by('-created_at').first()
        last_operation = AuditLog.objects.filter(user=user.username).select_related('product').order_by('-created_at').first()
        
        # العمليات الأخيرة (آخر 20 عملية)
        recent_operations = AuditLog.objects.filter(user=user.username).select_related('product').order_by('-created_at')[:20]
        
        # إحصائيات العمليات حسب النوع - استخدام values وannotate
        operation_counts = AuditLog.objects.filter(user=user.username).values('action').annotate(
            count=Count('id')
        )
        operation_stats = {
            dict(AuditLog.ACTION_CHOICES).get(item['action'], item['action']): item['count']
            for item in operation_counts
        }
        
        # إحصائيات الأنشطة للمسؤول - استخدام values وannotate
        activity_counts = UserActivityLog.objects.filter(
                user=user,
                created_at__gte=seven_days_ago
        ).values('action').annotate(count=Count('id'))
        
        activity_stats = {
            dict(UserActivityLog.ACTION_TYPES).get(item['action'], item['action']): item['count']
            for item in activity_counts
        }
        
        admin_members.append({
            'profile': profile,
            'user': user,
            'today_activities': activity_stats_agg['today_count'] or 0,
            'week_activities': activity_stats_agg['week_count'] or 0,
            'today_operations': operation_stats_agg['today_operations'] or 0,
            'week_operations': operation_stats_agg['week_operations'] or 0,
            'last_activity': last_activity,
            'last_operation': last_operation,
            'recent_operations': recent_operations,
            'operation_stats': operation_stats,
            'activity_stats': activity_stats,
            'last_login_ip': profile.last_login_ip,
        })
    
    # ترتيب المسؤولين حسب النشاط
    admin_members.sort(key=lambda x: x['today_activities'], reverse=True)
    
    # آخر 50 نشاط
    recent_logs = UserActivityLog.objects.select_related('user').order_by('-created_at')[:50]
    
    # إحصائيات الأنشطة حسب النوع - استخدام aggregation
    activity_counts = UserActivityLog.objects.filter(
            created_at__gte=seven_days_ago
    ).values('action').annotate(count=Count('id'))
    
    activity_stats = {
        dict(UserActivityLog.ACTION_TYPES).get(item['action'], item['action']): item['count']
        for item in activity_counts
    }
    
    context = {
        'total_staff': total_staff,
        'total_admins': total_admins,
        'recent_activities': recent_activities,
        'today_activities': today_activities,
        'staff_members': staff_members,
        'admin_members': admin_members,
        'recent_logs': recent_logs,
        'activity_stats': activity_stats,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=request.user,
        action='page_viewed',
        description='عرض لوحة تحكم المسؤول',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/admin_dashboard.html', context)


@staff_required
def staff_dashboard(request):
    """لوحة تحكم الموظف - إحصائيات شخصية"""
    user = request.user
    
    # إحصائيات اليوم
    today = timezone.now().date()
    today_activities = UserActivityLog.objects.filter(
        user=user,
        created_at__date=today
    ).count()
    
    # إحصائيات آخر 7 أيام
    seven_days_ago = timezone.now() - timedelta(days=7)
    week_activities = UserActivityLog.objects.filter(
        user=user,
        created_at__gte=seven_days_ago
    ).count()
    
    # آخر 20 نشاط
    recent_activities = UserActivityLog.objects.filter(user=user).order_by('-created_at')[:20]
    
    # الحصول على UserProfile
    user_profile = None
    if hasattr(user, 'user_profile'):
        user_profile = user.user_profile
    
    context = {
        'user': user,
        'user_profile': user_profile,
        'today_activities': today_activities,
        'week_activities': week_activities,
        'recent_activities': recent_activities,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=user,
        action='page_viewed',
        description='عرض لوحة تحكم الموظف',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/staff_dashboard.html', context)


@staff_required
def user_profile(request):
    """الملف الشخصي للمستخدم"""
    user = request.user
    
    # الحصول على UserProfile
    user_profile = None
    if hasattr(user, 'user_profile'):
        user_profile = user.user_profile
    else:
        # إنشاء profile إذا لم يكن موجوداً
        user_profile = UserProfile.objects.create(
            user=user,
            user_type='admin' if user.is_superuser else 'staff'
        )
    
    # جميع الأنشطة
    all_activities = UserActivityLog.objects.filter(user=user).order_by('-created_at')[:100]
    
    # إحصائيات حسب نوع النشاط
    activity_by_type = {}
    for action_code, action_name in UserActivityLog.ACTION_TYPES:
        count = UserActivityLog.objects.filter(user=user, action=action_code).count()
        if count > 0:
            activity_by_type[action_name] = count
    
    context = {
        'user': user,
        'user_profile': user_profile,
        'all_activities': all_activities,
        'activity_by_type': activity_by_type,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=user,
        action='page_viewed',
        description='عرض الملف الشخصي',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/user_profile.html', context)


@admin_required
def view_staff(request, user_id):
    """عرض تفاصيل موظف - للمسؤول فقط"""
    staff_user = get_object_or_404(User, id=user_id)
    staff_profile = get_object_or_404(UserProfile, user=staff_user)
    
    # جميع الأنشطة (UserActivityLog)
    all_activities = UserActivityLog.objects.filter(user=staff_user).order_by('-created_at')[:100]
    
    # جميع العمليات المفصلة (AuditLog) - عمليات المنتجات
    all_product_operations = AuditLog.objects.filter(user=staff_user.username).select_related('product').order_by('-created_at')[:200]
    
    # إحصائيات العمليات حسب النوع
    operation_stats = {}
    for action_code, action_name in AuditLog.ACTION_CHOICES:
        count = AuditLog.objects.filter(user=staff_user.username, action=action_code).count()
        if count > 0:
            operation_stats[action_name] = count
    
    # إحصائيات حسب نوع النشاط (UserActivityLog)
    activity_by_type = {}
    for action_code, action_name in UserActivityLog.ACTION_TYPES:
        count = UserActivityLog.objects.filter(user=staff_user, action=action_code).count()
        if count > 0:
            activity_by_type[action_name] = count
    
    # إحصائيات الوقت
    today = timezone.now().date()
    today_activities = UserActivityLog.objects.filter(user=staff_user, created_at__date=today).count()
    today_operations = AuditLog.objects.filter(user=staff_user.username, created_at__date=today).count()
    seven_days_ago = timezone.now() - timedelta(days=7)
    week_activities = UserActivityLog.objects.filter(user=staff_user, created_at__gte=seven_days_ago).count()
    week_operations = AuditLog.objects.filter(user=staff_user.username, created_at__gte=seven_days_ago).count()
    
    context = {
        'staff_user': staff_user,
        'staff_profile': staff_profile,
        'all_activities': all_activities,
        'all_product_operations': all_product_operations,
        'activity_by_type': activity_by_type,
        'operation_stats': operation_stats,
        'today_activities': today_activities,
        'today_operations': today_operations,
        'week_activities': week_activities,
        'week_operations': week_operations,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=request.user,
        action='user_viewed',
        description=f'عرض ملف الموظف {staff_user.username}',
        request=request,
        object_type='User',
        object_id=staff_user.id,
        object_name=staff_user.username
    )
    
    return render(request, 'inventory_app/view_staff.html', context)


@admin_required
@never_cache
@ratelimit(key='user', rate='20/h', method='POST', block=True) if RATELIMIT_AVAILABLE else lambda x: x
def edit_staff(request, user_id):
    """تعديل بيانات موظف - للمسؤول فقط مع تحسينات الأمان"""
    staff_user = get_object_or_404(User, id=user_id)
    staff_profile = get_object_or_404(UserProfile, user=staff_user)
    
    if request.method == 'POST':
        form = EditStaffForm(request.POST)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # تحديث بيانات المستخدم
                    username = form.cleaned_data['username']
                    email = form.cleaned_data['email']
                    
                    # التحقق من عدم تكرار اسم المستخدم
                    if username != staff_user.username:
                        if User.objects.filter(username=username).exclude(id=user_id).exists():
                            messages.error(request, 'اسم المستخدم مستخدم بالفعل')
                            return render(request, 'inventory_app/edit_staff.html', {
                                'staff_user': staff_user,
                                'staff_profile': staff_profile,
                                'form': form
                            })
                        staff_user.username = username
                    
                    staff_user.email = email if email else staff_user.email
                    
                    # تحديث كلمة السر إذا تم إدخالها
                    new_password = form.cleaned_data['password']
                    if new_password:
                        staff_user.set_password(new_password)
                        logger.info(f'Admin {request.user.username} changed password for user {staff_user.username}')
                    
                    staff_user.save()
                    
                    # تحديث UserProfile
                    staff_profile.phone = form.cleaned_data['phone']
                    staff_profile.notes = form.cleaned_data['notes']
                    staff_profile.user_type = form.cleaned_data['user_type']
                    staff_profile.save()
                
                # تسجيل النشاط
                UserActivityLog.log_activity(
                    user=request.user,
                    action='user_updated',
                    description=f'تعديل بيانات الموظف {staff_user.username}',
                    request=request,
                    object_type='User',
                    object_id=staff_user.id,
                    object_name=staff_user.username
                )
                
                logger.info(f'Admin {request.user.username} updated staff account: {staff_user.username}')
                messages.success(request, f'تم تحديث بيانات الموظف {staff_user.username} بنجاح')
                return redirect('inventory_app:admin_dashboard')
                
            except Exception as e:
                security_logger.error(f'Error updating staff account: {str(e)} from IP: {request.META.get("REMOTE_ADDR")}')
                logger.error(f'Error in edit_staff: {str(e)}')
                messages.error(request, f'حدث خطأ أثناء التحديث: {str(e)}')
        else:
            security_logger.warning(f'Invalid form data in edit_staff from IP: {request.META.get("REMOTE_ADDR")}')
            messages.error(request, 'يرجى التحقق من البيانات المدخلة:')
            for field, errors in form.errors.items():
                # Get field label if available
                field_label = field
                if field in form.fields:
                    field_label = form.fields[field].label or field
                    
                for error in errors:
                    messages.error(request, f"- {field_label}: {error}")
    else:
        form = EditStaffForm(initial={
            'username': staff_user.username,
            'email': staff_user.email,
            'phone': staff_profile.phone,
            'user_type': staff_profile.user_type,
            'notes': staff_profile.notes,
        })
    
    context = {
        'staff_user': staff_user,
        'staff_profile': staff_profile,
        'form': form,
    }
    
    return render(request, 'inventory_app/edit_staff.html', context)


def csrf_failure(request, reason=""):
    """معالجة أخطاء CSRF للأمان"""
    security_logger.warning(f'CSRF failure: {reason} from IP: {request.META.get("REMOTE_ADDR")} | User: {request.user.username if request.user.is_authenticated else "Anonymous"} | Path: {request.path}')
    
    messages.error(request, 'حدث خطأ أمني. يرجى المحاولة مرة أخرى.')
    return redirect('inventory_app:home')


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def toggle_staff_active(request, user_id):
    """تفعيل/تعطيل موظف - للمسؤول فقط"""
    try:
        staff_user = get_object_or_404(User, id=user_id)
        staff_profile = get_object_or_404(UserProfile, user=staff_user)
        
        # منع المسؤول من تعطيل نفسه
        if staff_user == request.user:
            return JsonResponse({
                'success': False,
                'error': 'لا يمكنك تعطيل نفسك'
            }, status=400)
        
        # تبديل الحالة
        staff_profile.is_active = not staff_profile.is_active
        staff_profile.save()
        
        action_text = 'تفعيل' if staff_profile.is_active else 'تعطيل'
        
        # تسجيل النشاط
        UserActivityLog.log_activity(
            user=request.user,
            action='user_updated',
            description=f'{action_text} الموظف {staff_user.username}',
            request=request,
            object_type='User',
            object_id=staff_user.id,
            object_name=staff_user.username,
            details={'is_active': staff_profile.is_active}
        )
        
        return JsonResponse({
            'success': True,
            'message': f'تم {action_text} الموظف {staff_user.username} بنجاح',
            'is_active': staff_profile.is_active
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ: {str(e)}'
        }, status=500)


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def delete_staff(request, user_id):
    """حذف موظف - للمسؤول فقط"""
    try:
        staff_user = get_object_or_404(User, id=user_id)
        staff_profile = get_object_or_404(UserProfile, user=staff_user)
        
        # منع المسؤول من حذف نفسه
        if staff_user == request.user:
            return JsonResponse({
                'success': False,
                'error': 'لا يمكنك حذف نفسك'
            }, status=400)
        
        username = staff_user.username
        
        # تسجيل النشاط قبل الحذف
        UserActivityLog.log_activity(
            user=request.user,
            action='user_deleted',
            description=f'حذف الموظف {username}',
            request=request,
            object_type='User',
            object_id=staff_user.id,
            object_name=username
        )
        
        # حذف UserProfile أولاً
        staff_profile.delete()
        # حذف User
        staff_user.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الموظف {username} بنجاح'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ: {str(e)}'
        }, status=500)



# ========== نظام المرتجعات ==========

@login_required
@staff_required
def returns_list(request):
    """عرض قائمة المرتجعات"""
    from django.core.paginator import Paginator
    
    # الاستعلام الأساسي
    returns_qs = ProductReturn.objects.all().order_by('-created_at')
    
    # إحصائيات
    total_returns = ProductReturn.objects.count()
    today_returns = ProductReturn.objects.filter(created_at__date=timezone.now().date()).count()
    total_quantities_returned = ProductReturn.objects.aggregate(
        total=db_models.Sum('total_quantities')
    )['total'] or 0
    
    # إضافة Pagination - 20 مرتجع لكل صفحة
    paginator = Paginator(returns_qs, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'returns': page_obj,
        'page_obj': page_obj,
        'total_returns': total_returns,
        'today_returns': today_returns,
        'total_quantities_returned': total_quantities_returned,
    }
    
    UserActivityLog.log_activity(
        user=request.user,
        action='page_viewed',
        description='عرض قائمة المرتجعات',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/returns_list.html', context)


@login_required
@staff_required
def add_return(request):
    """إضافة مرتجع جديد - صفحة النموذج"""
    products = Product.objects.select_related('location').all().order_by('product_number')
    
    context = {
        'products': products,
    }
    
    return render(request, 'inventory_app/add_return.html', context)


@login_required
@staff_required
@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def process_return(request):
    """معالجة المرتجع وإضافة الكميات للمنتجات - بدقة عالية جداً"""
    try:
        data = json.loads(request.body)
        products_list = data.get('products', [])
        return_reason = data.get('return_reason', '').strip()
        returned_by = data.get('returned_by', '').strip()
        notes = data.get('notes', '').strip()
        
        if not products_list:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم تحديد أي منتج للإرجاع'
            }, status=400)
        
        # التحقق من صحة البيانات والحسابات قبل المعالجة
        validated_products = []
        product_numbers = []
        
        for item in products_list:
            product_number = item.get('number', '').strip()
            quantity = item.get('quantity', 0)
            
            if not product_number:
                continue
            
            try:
                quantity = int(quantity)
                if quantity <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': f'كمية المنتج {product_number} يجب أن تكون أكبر من صفر'
                    }, status=400)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'error': f'كمية المنتج {product_number} غير صحيحة'
                }, status=400)
            
            product_numbers.append(product_number)
            validated_products.append({
                'product_number': product_number,
                'quantity': quantity
            })
        
        if not validated_products:
            return JsonResponse({
                'success': False,
                'error': 'لا توجد منتجات صحيحة للمعالجة'
            }, status=400)
        
        # الحصول على جميع المنتجات دفعة واحدة مع lock للتأكد من الدقة
        products_dict = {
            p.product_number: p 
            for p in Product.objects.filter(
                product_number__in=product_numbers
            ).select_for_update()
        }
        
        # التحقق من وجود جميع المنتجات
        missing_products = [p['product_number'] for p in validated_products if p['product_number'] not in products_dict]
        if missing_products:
            return JsonResponse({
                'success': False,
                'error': f'المنتجات التالية غير موجودة: {", ".join(missing_products)}'
            }, status=400)
        
        # معالجة المرتجع مع تسجيل دقيق
        return_products_data = []
        total_quantities = 0
        updated_products_count = 0
        
        for item in validated_products:
            product_number = item['product_number']
            return_quantity = item['quantity']
            
            product = products_dict[product_number]
            
            # الحصول على الكمية قبل التحديث بدقة
            old_quantity = product.quantity
            
            # إضافة الكمية المرتجعة بدقة عالية
            product.quantity = old_quantity + return_quantity
            
            # حفظ المنتج
            product.save(update_fields=['quantity'])
            
            # تسجيل في AuditLog بدقة
            AuditLog.objects.create(
                action='quantity_added',
                product=product,
                product_number=product_number,
                quantity_before=old_quantity,
                quantity_after=product.quantity,
                quantity_change=return_quantity,
                notes=f'إرجاع {return_quantity} من المرتجع',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            # إضافة إلى بيانات المرتجع
            return_products_data.append({
                'product_number': product_number,
                'product_name': product.name,
                'quantity_before': old_quantity,
                'quantity_returned': return_quantity,
                'quantity_after': product.quantity,
            })
            
            # حساب الإجماليات بدقة
            total_quantities += return_quantity
            updated_products_count += 1
        
        # إنشاء رقم مرتجع فريد
        from datetime import datetime
        import random
        import string
        
        return_number = f"RET-{datetime.now().strftime('%Y%m%d%H%M%S')}-{''.join(random.choices(string.ascii_uppercase + string.digits, k=4))}"
        
        # إنشاء المرتجع
        product_return = ProductReturn.objects.create(
            return_number=return_number,
            products_data=return_products_data,
            total_products=updated_products_count,
            total_quantities=total_quantities,
            return_reason=return_reason if return_reason else None,
            returned_by=returned_by if returned_by else None,
            notes=notes if notes else None,
            user=request.user.username if request.user.is_authenticated else 'Guest'
        )
        
        # تسجيل النشاط
        UserActivityLog.log_activity(
            user=request.user,
            action='order_created',
            description=f'تم إنشاء مرتجع جديد: {return_number} - عدد المنتجات: {updated_products_count} - الكمية: {total_quantities}',
            request=request,
            object_type='ProductReturn',
            object_id=product_return.id,
            object_name=return_number
        )
        
        logger.info(f'Return created: {return_number} by {request.user.username}, Products: {updated_products_count}, Quantities: {total_quantities}')
        
        return JsonResponse({
            'success': True,
            'message': f'تم إضافة المرتجع بنجاح - تم إضافة {total_quantities} كمية إلى {updated_products_count} منتج',
            'return_number': return_number,
            'return_id': product_return.id,
            'products_updated': updated_products_count,
            'total_quantities': total_quantities,
            'return_data': return_products_data
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'خطأ في قراءة البيانات'
        }, status=400)
    except Exception as e:
        logger.error(f'Error processing return: {str(e)}')
        security_logger.error(f'Error processing return: {str(e)} from IP: {request.META.get("REMOTE_ADDR")}')
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ أثناء معالجة المرتجع: {str(e)}'
        }, status=500)


@login_required
@staff_required
def return_detail(request, return_id):
    """عرض تفاصيل مرتجع معين"""
    product_return = get_object_or_404(ProductReturn, id=return_id)
    
    context = {
        'return': product_return,
    }
    
    UserActivityLog.log_activity(
        user=request.user,
        action='page_viewed',
        description=f'عرض تفاصيل المرتجع {product_return.return_number}',
        request=request,
        object_type='ProductReturn',
        object_id=product_return.id,
        object_name=product_return.return_number
    )
    
    return render(request, 'inventory_app/return_detail.html', context)


# ==================== استيراد المنتجات من Excel ====================

@login_required
@admin_required
def import_products_excel(request):
    """صفحة استيراد المنتجات من ملف Excel"""
    return render(request, 'inventory_app/import_excel.html')


@login_required
@admin_required
def upload_excel_file(request):
    """رفع ومعالجة ملف Excel"""
    if request.method != 'POST':
        return JsonResponse({'error': 'طريقة غير مسموحة'}, status=405)
    
    if 'excel_file' not in request.FILES:
        return JsonResponse({'error': 'لم يتم رفع ملف'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # التحقق من نوع الملف
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'error': 'يجب أن يكون الملف بصيغة Excel (.xlsx أو .xls)'}, status=400)
    
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        # قراءة الملف
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # قراءة البيانات
        data = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                # الصف الأول = العناوين
                headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row, start=1)]
            else:
                # البيانات
                if any(cell is not None for cell in row):  # تجاهل الصفوف الفارغة
                    data.append(list(row))
        
        request.session['excel_data'] = {
            'headers': headers,
            'data': data
        }
        
        total_rows = len(data)
        # تم تحديث الحد المسموح به للمعاينة ليشمل كامل الملف
        # هذا يضمن أن المستخدم يرى ويعالج جميع البيانات، خاصة للملفات المتوسطة الحجم (مثل 420 منتج)
        preview_limit = total_rows
        if total_rows > 5000:
            # تنبيه: للملفات الكبيرة جداً، قد يكون الأداء بطيئاً في المتصفح
            # لكننا سنسمح بذلك لضمان الدقة
            preview_limit = total_rows
        
        return JsonResponse({
            'success': True,
            'headers': headers,
            'data': data[:min(preview_limit, 50)],
            'total_rows': total_rows,
            'preview_limit': preview_limit
        })
        
    except Exception as e:
        return JsonResponse({'error': f'خطأ في قراءة الملف: {str(e)}'}, status=500)


@login_required
@admin_required
def preview_excel_data(request):
    """معاينة البيانات من Excel قبل الإضافة"""
    if request.method != 'POST':
        return JsonResponse({'error': 'طريقة غير مسموحة'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        # الحصول على البيانات من الجلسة
        excel_data = request.session.get('excel_data')
        if not excel_data:
            return JsonResponse({'error': 'لم يتم العثور على بيانات Excel'}, status=400)
        
        # الحصول على تحديد الأعمدة
        column_mapping = data.get('column_mapping', {})
        
        # التحقق من الأعمدة المطلوبة
        required_fields = ['product_number', 'total_quantity']
        for field in required_fields:
            if field not in column_mapping or column_mapping[field] is None:
                return JsonResponse({'error': f'يجب تحديد عمود {field}'}, status=400)
        
        headers = excel_data['headers']
        total_rows = len(excel_data['data'])
        
        # السماح بمعاينة كافة الصفوف
        preview_limit = total_rows
        
        rows = excel_data['data'][:preview_limit]
        
        # التحقق من وجود أعمدة اختيارية
        has_final_model = 'final_model' in column_mapping and column_mapping['final_model'] is not None
        has_name = 'name' in column_mapping and column_mapping['name'] is not None
        has_category = 'category' in column_mapping and column_mapping['category'] is not None
        
        print(f"[DEBUG] Total rows in Excel: {len(rows)}")
        print(f"[DEBUG] Has FINAL_MODEL column: {has_final_model}")
        
        # معاينة البيانات
        preview_data = []
        
        for row_idx, row in enumerate(rows, start=2):
            try:
                # التحقق من أن الصف ليس فارغاً
                if not any(cell for cell in row if cell is not None):
                    continue
                
                # استخراج البيانات
                try:
                    product_number_cell = row[column_mapping['product_number']]
                    product_number = str(product_number_cell).strip() if product_number_cell is not None else None
                    
                    # تخطي صفوف العناوين
                    if product_number and product_number.upper() in ['MODEL', 'PRODUCT', 'رقم المنتج', 'PRODUCT NUMBER']:
                        continue
                except (IndexError, KeyError):
                    product_number = None
                
                # استخراج FINAL_MODEL إذا كان موجوداً
                final_model = None
                if has_final_model:
                    try:
                        final_model_cell = row[column_mapping['final_model']]
                        final_model = str(final_model_cell).strip() if final_model_cell is not None else None
                        
                        # تخطي صفوف العناوين
                        if final_model and final_model.upper() in ['FINAL MODEL', 'FINAL_MODEL', 'رقم المنتج النهائي']:
                            continue
                    except (IndexError, KeyError):
                        final_model = None
                
                # إجمالي الكمية (مطلوب)
                try:
                    total_cell = row[column_mapping['total_quantity']]
                    if total_cell and isinstance(total_cell, (int, float)):
                        total_quantity = int(total_cell)
                    elif total_cell and str(total_cell).replace('.', '').replace('-', '').isdigit():
                        total_quantity = int(float(total_cell))
                    else:
                        total_quantity = None
                except (IndexError, KeyError, ValueError):
                    total_quantity = None
                
                # حقول اختيارية: الاسم والفئة
                name_val = None
                category_val = None
                try:
                    if 'name' in column_mapping and column_mapping['name'] is not None:
                        nc = row[column_mapping['name']]
                        name_val = str(nc).strip() if nc is not None else None
                except Exception:
                    name_val = None
                try:
                    if 'category' in column_mapping and column_mapping['category'] is not None:
                        cc = row[column_mapping['category']]
                        category_val = str(cc).strip() if cc is not None else None
                except Exception:
                    category_val = None
                
                # الموقع
                location_str = None
                try:
                    if 'location' in column_mapping and column_mapping['location'] is not None:
                        loc_cell = row[column_mapping['location']]
                        location_str = str(loc_cell).strip() if loc_cell else None
                except (IndexError, KeyError):
                    location_str = None
                
                # تحديد رقم المنتج النهائي
                # إذا كان هناك عمود FINAL_MODEL، استخدمه مباشرة
                if has_final_model and final_model:
                    final_product_number = final_model
                else:
                    # إذا لم يكن هناك FINAL_MODEL، استخدم product_number
                    final_product_number = product_number
                
                # التحقق من البيانات
                if not final_product_number or total_quantity is None:
                    error_details = []
                    if not final_product_number:
                        error_details.append('رقم المنتج فارغ')
                    if total_quantity is None:
                        error_details.append('الإجمالي فارغ')
                    
                    preview_data.append({
                        'row': row_idx,
                        'original_number': str(product_number_cell).strip() if product_number_cell and str(product_number_cell).strip() else 'فارغ',
                        'final_number': final_product_number or 'فارغ',
                        'total_quantity': total_quantity or 0,
                        'location': location_str or '',
                        'status': 'error',
                        'message': f'⚠️ {" | ".join(error_details)}'
                    })
                    continue
                
                # التحقق من وجود المنتج
                existing_product = Product.objects.filter(product_number=final_product_number).first()
                
                status = 'new'
                message = 'منتج جديد'
                
                if existing_product:
                    status = 'exists'
                    message = f'موجود مسبقاً (الكمية الحالية: {existing_product.quantity})'
                
                result_row = {
                    'row': row_idx,
                    'original_number': str(product_number_cell).strip() if product_number_cell else product_number,
                    'final_number': final_product_number,
                    'total_quantity': total_quantity,
                    'location': location_str or '',
                    'status': status,
                    'message': message
                }
                if has_name:
                    try:
                        name_cell = row[column_mapping['name']]
                        result_row['name'] = str(name_cell).strip() if name_cell is not None else ''
                    except Exception:
                        result_row['name'] = ''
                if has_category:
                    try:
                        category_cell = row[column_mapping['category']]
                        result_row['category'] = str(category_cell).strip() if category_cell is not None else ''
                    except Exception:
                        result_row['category'] = ''
                preview_data.append(result_row)
                
            except Exception as e:
                preview_data.append({
                    'row': row_idx,
                    'original_number': 'N/A',
                    'final_number': 'N/A',
                    'total_quantity': 0,
                    'location': '',
                    'status': 'error',
                    'message': str(e)
                })
        
        print(f"[DEBUG] Preview data count: {len(preview_data)}")
        
        return JsonResponse({
            'success': True,
            'preview': preview_data,
            'total_rows': len(preview_data),
            'preview_limit': preview_limit,
            'full_rows': total_rows
        })
        
    except Exception as e:
        return JsonResponse({'error': f'خطأ في المعاينة: {str(e)}'}, status=500)


@login_required
@admin_required
@transaction.atomic
def process_excel_data(request):
    """معالجة البيانات من Excel وإضافتها للنظام"""
    if request.method != 'POST':
        return JsonResponse({'error': 'طريقة غير مسموحة'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        # الحصول على البيانات من الجلسة
        excel_data = request.session.get('excel_data')
        if not excel_data:
            return JsonResponse({'error': 'لم يتم العثور على بيانات Excel'}, status=400)
        
        # الحصول على تحديد الأعمدة
        column_mapping = data.get('column_mapping', {})
        conflict_resolution = data.get('conflict_resolution', 'skip')  # skip, update, replace
        edited_data = data.get('edited_data', None)  # البيانات المعدلة من المعاينة
        
        # التحقق من الأعمدة المطلوبة
        required_fields = ['product_number', 'total_quantity']
        for field in required_fields:
            if field not in column_mapping or column_mapping[field] is None:
                return JsonResponse({'error': f'يجب تحديد عمود {field}'}, status=400)
        
        headers = excel_data['headers']
        rows = excel_data['data']
        
        print(f"[DEBUG] Edited data received: {edited_data is not None}")
        if edited_data:
            print(f"[DEBUG] Edited data count: {len(edited_data)}")
        
        # معالجة البيانات
        results = {
            'added': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }
        
        from django.db import transaction
        
        # تتبع تكرار أرقام المنتجات لإضافة -1, -2, etc
        product_number_counter = {}
        
        # إذا كانت هناك بيانات معدلة، استخدمها بدلاً من قراءة Excel مرة أخرى
        has_name = 'name' in column_mapping and column_mapping['name'] is not None
        has_category = 'category' in column_mapping and column_mapping['category'] is not None
        if edited_data:
            print("[DEBUG] Using edited data from preview")
            for item in edited_data:
                # تخطي الصفوف التي ما زالت تحتوي على أخطاء
                if item['status'] == 'error':
                    results['errors'].append(f"الصف {item['row']}: {item['message']}")
                    results['skipped'] += 1
                    continue
                
                try:
                    product_number = item['final_number']
                    location_str = item.get('location', None)
                    # استخراج الاسم والفئة إذا طُلبا
                    name_val = None
                    category_val = None
                    try:
                        src_row = rows[item['row'] - 2]
                        if has_name and column_mapping['name'] is not None:
                            nc = src_row[column_mapping['name']]
                            name_val = str(nc).strip() if nc is not None else None
                        if has_category and column_mapping['category'] is not None:
                            cc = src_row[column_mapping['category']]
                            category_val = str(cc).strip() if cc is not None else None
                    except Exception:
                        pass
                    
                    # إجمالي الكمية مباشرة
                    total_quantity = item['total_quantity']
                    
                    # معالجة الموقع
                    location = None
                    if location_str:
                        # دعم صيغ متعددة مثل "R1C5" أو "R1-C5" أو "R1 C5"
                        import re
                        s = str(location_str).strip().upper()
                        # تحويل أي فاصل غير حرفي إلى صيغة R..C..
                        s = s.replace('-', ' ').replace('_', ' ')
                        s = re.sub(r"\s+", "", s)
                        match = re.match(r"R(\d+)C(\d+)", s)
                        if match:
                            row_num = int(match.group(1))
                            col_num = int(match.group(2))
                            warehouse = Warehouse.objects.first()
                            if warehouse:
                                location, _ = Location.objects.get_or_create(
                                    warehouse=warehouse,
                                    row=row_num,
                                    column=col_num
                                )
                        else:
                            # كاحتياط: إذا كانت القيمة رقمية فقط فتعامل معها كمُعرّف موقع
                            if s.isdigit():
                                try:
                                    location = Location.objects.get(id=int(s))
                                except Location.DoesNotExist:
                                    location = None
                    
                    # التحقق من وجود المنتج
                    existing_product = Product.objects.filter(product_number=product_number).first()
                    
                    if existing_product:
                        if conflict_resolution == 'skip':
                            results['skipped'] += 1
                            continue
                        elif conflict_resolution == 'update':
                            existing_product.quantity += total_quantity
                            if location:
                                existing_product.location = location
                            update_fields = ['quantity', 'location']
                            if name_val:
                                existing_product.name = name_val
                                update_fields.append('name')
                            if category_val:
                                existing_product.category = category_val
                                update_fields.append('category')
                            existing_product.save(update_fields=update_fields)
                            results['updated'] += 1
                        elif conflict_resolution == 'replace':
                            existing_product.quantity = total_quantity
                            if location:
                                existing_product.location = location
                            update_fields = ['quantity', 'location']
                            if name_val:
                                existing_product.name = name_val
                                update_fields.append('name')
                            if category_val:
                                existing_product.category = category_val
                                update_fields.append('category')
                            existing_product.save(update_fields=update_fields)
                            results['updated'] += 1
                    else:
                        # إنشاء منتج جديد
                        Product.objects.create(
                            product_number=product_number,
                            name=(name_val or product_number),
                            category=(category_val or None),
                            quantity=total_quantity,
                            location=location
                        )
                        results['added'] += 1
                        
                except Exception as e:
                    results['errors'].append(f"الصف {item['row']}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'results': results
            })
        
        # إذا لم تكن هناك بيانات معدلة، استخدم القراءة المباشرة من Excel
        has_final_model = 'final_model' in column_mapping and column_mapping['final_model'] is not None
        
        for row_idx, row in enumerate(rows, start=2):  # start=2 لأن الصف 1 هو العناوين
            try:
                # التحقق من أن الصف ليس فارغاً تماماً
                if not any(cell for cell in row if cell is not None):
                    continue
                
                # استخراج البيانات مع معالجة الأخطاء
                try:
                    product_number_cell = row[column_mapping['product_number']]
                    product_number = str(product_number_cell).strip() if product_number_cell is not None else None
                    
                    # تخطي صفوف العناوين المكررة
                    if product_number and product_number.upper() in ['MODEL', 'PRODUCT', 'رقم المنتج', 'PRODUCT NUMBER']:
                        continue
                except (IndexError, KeyError):
                    product_number = None
                
                # استخراج FINAL_MODEL إذا كان موجوداً
                final_model = None
                if has_final_model:
                    try:
                        final_model_cell = row[column_mapping['final_model']]
                        final_model = str(final_model_cell).strip() if final_model_cell is not None else None
                        
                        if final_model and final_model.upper() in ['FINAL MODEL', 'FINAL_MODEL', 'رقم المنتج النهائي']:
                            continue
                    except (IndexError, KeyError):
                        final_model = None
                
                # إجمالي الكمية (مطلوب)
                try:
                    total_cell = row[column_mapping['total_quantity']]
                    if total_cell and isinstance(total_cell, (int, float)):
                        total_quantity = int(total_cell)
                    elif total_cell and str(total_cell).replace('.', '').replace('-', '').isdigit():
                        total_quantity = int(float(total_cell))
                    else:
                        total_quantity = None
                except (IndexError, KeyError, ValueError):
                    total_quantity = None
                
                # حقول اختيارية
                location_str = None
                try:
                    if 'location' in column_mapping and column_mapping['location'] is not None:
                        loc_cell = row[column_mapping['location']]
                        location_str = str(loc_cell).strip() if loc_cell else None
                except (IndexError, KeyError):
                    location_str = None
                
                # حقول اختيارية: الاسم والفئة
                name_val = None
                category_val = None
                try:
                    if 'name' in column_mapping and column_mapping['name'] is not None:
                        nc = row[column_mapping['name']]
                        name_val = str(nc).strip() if nc is not None else None
                except Exception:
                    name_val = None
                try:
                    if 'category' in column_mapping and column_mapping['category'] is not None:
                        cc = row[column_mapping['category']]
                        category_val = str(cc).strip() if cc is not None else None
                except Exception:
                    category_val = None
                
                # تحديد رقم المنتج النهائي
                if has_final_model and final_model:
                    final_product_number = final_model
                else:
                    final_product_number = product_number
                
                # التحقق من البيانات
                if not final_product_number or total_quantity is None:
                    results['errors'].append(f'الصف {row_idx}: بيانات ناقصة (رقم المنتج أو الإجمالي)')
                    continue
                
                # البحث عن الموقع
                location = None
                if location_str:
                    # تحليل الموقع (مثل R1C5)
                    import re
                    match = re.match(r'R(\d+)C(\d+)', location_str.upper())
                    if match:
                        row_num = int(match.group(1))
                        col_num = int(match.group(2))
                        
                        # البحث عن الموقع أو إنشاءه
                        warehouse = Warehouse.objects.first()
                        if warehouse:
                            location, _ = Location.objects.get_or_create(
                                warehouse=warehouse,
                                row=row_num,
                                column=col_num
                            )
                
                # التحقق من وجود المنتج
                existing_product = Product.objects.filter(product_number=final_product_number).first()
                
                if existing_product:
                    # المنتج موجود - معالجة التعارض
                    if conflict_resolution == 'skip':
                        results['skipped'] += 1
                    elif conflict_resolution == 'update':
                        # تحديث الكمية (إضافة)
                        existing_product.quantity += total_quantity
                        if location:
                            existing_product.location = location
                        if name_val:
                            existing_product.name = name_val
                        if category_val:
                            existing_product.category = category_val
                        existing_product.save()
                        results['updated'] += 1
                    elif conflict_resolution == 'replace':
                        # استبدال الكمية
                        existing_product.quantity = total_quantity
                        if location:
                            existing_product.location = location
                        if name_val:
                            existing_product.name = name_val
                        if category_val:
                            existing_product.category = category_val
                        existing_product.save()
                        results['updated'] += 1
                else:
                    # منتج جديد
                    Product.objects.create(
                        product_number=final_product_number,
                        name=(name_val or final_product_number),
                        category=(category_val or None),
                        quantity=total_quantity,
                        location=location
                    )
                    results['added'] += 1
                    
            except Exception as e:
                results['errors'].append(f'الصف {row_idx}: {str(e)}')
        
        # مسح البيانات من الجلسة
        if 'excel_data' in request.session:
            del request.session['excel_data']
        
        return JsonResponse({
            'success': True,
            'results': results
        })
        
    except Exception as e:
        return JsonResponse({'error': f'خطأ في معالجة البيانات: {str(e)}'}, status=500)


# ==================== دمج ملفات متعددة (Excel/JSON) ====================

@login_required
@admin_required
def merge_files_page(request):
    return render(request, 'inventory_app/merge_files.html')


def _normalize_product_number(raw):
    import re
    if raw is None:
        return ''
    s = str(raw).strip()
    return re.sub(r'[^A-Za-z0-9]', '', s).upper()


def _extract_products_from_excel(file_obj):
    from openpyxl import load_workbook
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers = []
    products = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(c).strip().lower() if c is not None else '' for c in row]
            continue
        # heuristic mapping
        # find product_number column
        def _norm(s):
            s = (s or '').strip().lower()
            # إزالة المسافات وبعض العلامات فقط مع إبقاء الأحرف العربية
            return ''.join(ch for ch in s if ch not in ' .,_-()[]{}\n\t')
        def find_col(names):
            norm_headers = [_norm(h) for h in headers]
            for n in names:
                nn = _norm(n)
                # مطابقة كاملة أو تحتوي
                for i, nh in enumerate(norm_headers):
                    if nh == nn or (nn and nn in nh):
                        return i
            return None
        pn_idx = find_col(['final_model', 'final model', 'model', 'product number', 'productnumber', 'رقم المنتج النهائي', 'رقم المنتج', 'موديل', 'الموديل'])
        qty_idx = find_col(['total_quantity', 'total qty', 'total', 'quantity', 'qty', 't.qty', 'tqty', 'الاجمالي', 'اجمالي', 'الكمية', 'كمية'])
        if qty_idx is None:
            for i, h in enumerate(headers):
                hn = _norm(h)
                if 'qty' in hn or 'quantity' in hn or 'الكمية' in h or 'اجمالي' in h or 'الاجمالي' in h:
                    qty_idx = i
                    break
        name_idx = find_col(['name', 'الاسم'])
        if pn_idx is None and len(row) > 0:
            # fallback: first non-empty cell treated as product number
            try:
                pn_idx = next(i for i, c in enumerate(row) if c is not None)
            except StopIteration:
                pn_idx = None
        if pn_idx is None:
            continue
        product_number = str(row[pn_idx]).strip() if row[pn_idx] is not None else ''
        name = str(row[name_idx]).strip() if (name_idx is not None and row[name_idx] is not None) else product_number
        total_quantity = 0
        if qty_idx is not None:
            q = row[qty_idx]
            if isinstance(q, (int, float)):
                total_quantity = int(q)
            elif q and str(q).replace('.', '').replace('-', '').isdigit():
                try:
                    total_quantity = int(float(q))
                except Exception:
                    total_quantity = 0
        products.append({'product_number': product_number, 'name': name, 'quantity': total_quantity})
    return products


def _extract_products_from_json(file_obj):
    import json as pyjson
    text = file_obj.read().decode('utf-8') if hasattr(file_obj, 'read') else file_obj.decode('utf-8')
    data = pyjson.loads(text)
    products = []
    # backup schema
    if isinstance(data, dict) and 'products' in data and isinstance(data['products'], list):
        for rec in data['products']:
            if isinstance(rec, dict):
                fields = rec.get('fields', {}) if 'fields' in rec else rec
                pn = fields.get('product_number') or fields.get('final_model') or ''
                name = fields.get('name') or pn
                qty = fields.get('quantity') or 0
                products.append({'product_number': pn, 'name': name, 'quantity': int(qty) if isinstance(qty, (int, float)) else 0})
    elif isinstance(data, list):
        for rec in data:
            pn = rec.get('product_number') or rec.get('final_model') or ''
            name = rec.get('name') or pn
            qty = rec.get('quantity') or 0
            products.append({'product_number': pn, 'name': name, 'quantity': int(qty) if isinstance(qty, (int, float)) else 0})
    return products


@login_required
@admin_required
@require_http_methods(["POST"])
def merge_files_upload(request):
    try:
        files = request.FILES.getlist('files[]') or request.FILES.getlist('files')
        if not files:
            return JsonResponse({'success': False, 'error': 'يرجى اختيار ملفات'}, status=400)
        all_items = []
        for f in files:
            fname = f.name.lower()
            items = []
            if fname.endswith(('.xlsx', '.xls')):
                items = _extract_products_from_excel(f)
            elif fname.endswith('.json'):
                items = _extract_products_from_json(f)
            else:
                continue
            for it in items:
                it['source'] = f.name
            all_items.extend(items)
        # duplicates detection
        groups = {}
        for it in all_items:
            key = _normalize_product_number(it.get('product_number'))
            if not key:
                continue
            groups.setdefault(key, []).append(it)
        duplicates = {k: v for k, v in groups.items() if len(v) > 1}
        return JsonResponse({
            'success': True,
            'items': all_items,
            'counts': {
                'total': len(all_items),
                'unique': len(groups),
                'duplicates': len(duplicates)
            },
            'duplicates': duplicates
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        logger.error(f'Merge upload error: {str(e)}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@admin_required
@require_http_methods(["POST"])
def merge_files_process(request):
    try:
        import json as pyjson
        body = pyjson.loads(request.body)
        items = body.get('items', [])
        auto_fix = body.get('auto_fix', 'merge')  # merge, rename
        # consolidate
        consolidated = {}
        for it in items:
            key = _normalize_product_number(it.get('product_number'))
            if not key:
                continue
            if key not in consolidated:
                consolidated[key] = {
                    'product_number': it.get('product_number') or '',
                    'name': it.get('name') or it.get('product_number') or '',
                    'quantity': int(it.get('quantity') or 0)
                }
            else:
                if auto_fix == 'merge':
                    consolidated[key]['quantity'] += int(it.get('quantity') or 0)
                elif auto_fix == 'rename':
                    # keep separate by appending suffix
                    # store as list of variants
                    consolidated.setdefault('_variants', []).append(it)
        result_list = list(consolidated.values())
        # if rename, append variants with unique suffixes
        if auto_fix == 'rename' and '_variants' in consolidated:
            idx = 1
            for it in consolidated['_variants']:
                base = str(it.get('product_number') or '')
                result_list.append({
                    'product_number': f"{base}-{idx}",
                    'name': it.get('name') or base,
                    'quantity': int(it.get('quantity') or 0)
                })
                idx += 1
        return JsonResponse({'success': True, 'items': result_list, 'total': len(result_list)})
    except Exception as e:
        logger.error(f'Merge process error: {str(e)}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@admin_required
@require_http_methods(["POST"])
def merge_files_export(request):
    try:
        import json as pyjson
        export_format = request.POST.get('format', 'json')
        items_json = request.POST.get('items')
        items = pyjson.loads(items_json) if items_json else []
        from django.http import HttpResponse
        if export_format == 'excel':
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'المنتجات'
            ws.append(['رقم المنتج', 'الاسم', 'الكمية'])
            for it in items:
                ws.append([it.get('product_number') or '', it.get('name') or '', int(it.get('quantity') or 0)])
            resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = 'attachment; filename="merged_products.xlsx"'
            wb.save(resp)
            return resp
        else:
            resp = HttpResponse(pyjson.dumps(items, ensure_ascii=False), content_type='application/json')
            resp['Content-Disposition'] = 'attachment; filename="merged_products.json"'
            return resp
    except Exception as e:
        logger.error(f'Merge export error: {str(e)}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ==================== Container Management ====================

@login_required
@staff_required
def container_list(request):
    """عرض قائمة الحاويات"""
    containers = Container.objects.all().annotate(
        products_count=db_models.Count('products')
    )
    
    context = {
        'containers': containers,
        'total_containers': containers.count()
    }
    
    return render(request, 'inventory_app/container_list.html', context)


@login_required
@staff_required
@require_http_methods(["POST"])
def container_add(request):
    """إضافة حاوية جديدة"""
    try:
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        color = request.POST.get('color', '#667eea')
        
        if not name:
            return JsonResponse({
                'success': False,
                'error': 'اسم الحاوية مطلوب'
            }, status=400)
        
        # التحقق من عدم وجود حاوية بنفس الاسم
        if Container.objects.filter(name=name).exists():
            return JsonResponse({
                'success': False,
                'error': 'يوجد حاوية بنفس الاسم بالفعل'
            }, status=400)
        
        container = Container.objects.create(
            name=name,
            description=description,
            color=color
        )
        
        return JsonResponse({
            'success': True,
            'message': 'تم إضافة الحاوية بنجاح',
            'container': {
                'id': container.id,
                'name': container.name,
                'color': container.color
            }
        })
        
    except Exception as e:
        logger.error(f'Error adding container: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@staff_required
@require_http_methods(["POST"])
def assign_products_to_container(request):
    """تعيين منتجات إلى حاوية"""
    try:
        data = json.loads(request.body)
        product_ids = data.get('product_ids', [])
        container_id = data.get('container_id')
        
        if not product_ids:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم تحديد أي منتجات'
            }, status=400)
        
        # إذا كان container_id هو None أو "null"، نقوم بإزالة المنتجات من الحاوية
        if container_id in [None, 'null', '']:
            Product.objects.filter(id__in=product_ids).update(container=None)
            return JsonResponse({
                'success': True,
                'message': f'تم إزالة {len(product_ids)} منتج من الحاوية'
            })
        
        # التحقق من وجود الحاوية
        container = get_object_or_404(Container, id=container_id)
        
        # تعيين المنتجات للحاوية
        updated_count = Product.objects.filter(id__in=product_ids).update(container=container)
        
        return JsonResponse({
            'success': True,
            'message': f'تم تعيين {updated_count} منتج إلى الحاوية "{container.name}"'
        })
        
    except Container.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'الحاوية غير موجودة'
        }, status=404)
    except Exception as e:
        logger.error(f'Error assigning products to container: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@staff_required
@require_http_methods(["POST"])
def container_delete(request, container_id):
    """حذف حاوية"""
    try:
        container = get_object_or_404(Container, id=container_id)
        
        # إزالة المنتجات من الحاوية قبل الحذف
        Product.objects.filter(container=container).update(container=None)
        
        container_name = container.name
        container.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الحاوية "{container_name}" بنجاح'
        })
        
    except Container.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'الحاوية غير موجودة'
        }, status=404)
    except Exception as e:
        logger.error(f'Error deleting container: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
def _load_backup_data(raw_bytes, filename):
    data = None
    parse_info = {'source': 'raw', 'encoding': None, 'fixes': []}
    try:
        if raw_bytes[:2] == b'PK' or (filename or '').lower().endswith('.zip'):
            import zipfile, io
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
                json_members = [m for m in zf.namelist() if m.lower().endswith('.json')]
                if not json_members:
                    return None, {'error': 'zip_no_json'}
                raw_bytes = zf.read(json_members[0])
                parse_info['source'] = 'zip'
    except Exception:
        pass
    import codecs, re
    # BOM detection (wrapped)
    try:
        if raw_bytes.startswith(codecs.BOM_UTF8):
            text = raw_bytes.decode('utf-8', errors='ignore')
            parse_info['encoding'] = 'utf-8-bom'
            t = re.sub(r"//.*?$", "", text, flags=re.MULTILINE)
            t = re.sub(r"/\*.*?\*/", "", t, flags=re.DOTALL)
            t = re.sub(r",\s*([}\]])", r"\1", t)
            try:
                return json.loads(t), parse_info
            except Exception:
                pass
        elif raw_bytes.startswith(codecs.BOM_UTF16_LE):
            text = raw_bytes.decode('utf-16-le', errors='ignore')
            parse_info['encoding'] = 'utf-16-le-bom'
            parse_info['fixes'].append('ignore_errors')
            t = re.sub(r"//.*?$", "", text, flags=re.MULTILINE)
            t = re.sub(r"/\*.*?\*/", "", t, flags=re.DOTALL)
            t = re.sub(r",\s*([}\]])", r"\1", t)
            try:
                return json.loads(t), parse_info
            except Exception:
                pass
        elif raw_bytes.startswith(codecs.BOM_UTF16_BE):
            text = raw_bytes.decode('utf-16-be', errors='ignore')
            parse_info['encoding'] = 'utf-16-be-bom'
            parse_info['fixes'].append('ignore_errors')
            t = re.sub(r"//.*?$", "", text, flags=re.MULTILINE)
            t = re.sub(r"/\*.*?\*/", "", t, flags=re.DOTALL)
            t = re.sub(r",\s*([}\]])", r"\1", t)
            try:
                return json.loads(t), parse_info
            except Exception:
                pass
    except Exception:
        pass
    # Try encodings list
    encodings = ['utf-8', 'utf-8-sig', 'cp1256', 'windows-1256', 'latin1', 'utf-16', 'utf-16le', 'utf-16be', 'utf-32', 'utf-32le', 'utf-32be']
    last_err = None
    for enc in encodings:
        for err_mode in ['strict', 'ignore']:
            try:
                text = raw_bytes.decode(enc, errors=err_mode)
                parse_info['encoding'] = enc if err_mode == 'strict' else f'{enc}-ignore'
                # Clean and slice
                t = text.lstrip('\ufeff')
                t = re.sub(r"//.*?$", "", t, flags=re.MULTILINE)
                t = re.sub(r"/\*.*?\*/", "", t, flags=re.DOTALL)
                t = re.sub(r",\s*([}\]])", r"\1", t)
                try:
                    if err_mode == 'ignore':
                        parse_info['fixes'].append('ignore_errors')
                    parse_info['fixes'].append('clean')
                    return json.loads(t), parse_info
                except Exception as e2:
                    last_err = str(e2)
                    try:
                        s = t.find('{')
                        epos = t.rfind('}')
                        if s != -1 and epos != -1 and epos > s:
                            t2 = t[s:epos+1]
                            parse_info['fixes'].append('slice')
                            return json.loads(t2), parse_info
                    except Exception as e3:
                        last_err = str(e3)
            except Exception as e:
                last_err = str(e)
                continue
    # Binary brace slice fallback
    try:
        s = raw_bytes.find(b'{')
        epos = raw_bytes.rfind(b'}')
        if s != -1 and epos != -1 and epos > s:
            segment = raw_bytes[s:epos+1]
            for enc in encodings:
                try:
                    t = segment.decode(enc, errors='ignore')
                    t = re.sub(r"//.*?$", "", t, flags=re.MULTILINE)
                    t = re.sub(r"/\*.*?\*/", "", t, flags=re.DOTALL)
                    t = re.sub(r",\s*([}\]])", r"\1", t)
                    parse_info['encoding'] = f'{enc}-segment-ignore'
                    parse_info['fixes'].append('binary_slice')
                    return json.loads(t), parse_info
                except Exception:
                    continue
    except Exception as e:
        last_err = str(e)
    return None, {'error': 'json_invalid', 'message': last_err}


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def compact_row(request):
    """
    إعادة ترتيب المنتجات في صف معين لملء الفراغات.
    يتم نقل المنتجات من الخلايا البعيدة إلى الخلايا الفارغة الأقرب للبداية (العمود 1).
    """
    try:
        data = json.loads(request.body)
        row_number = int(data.get('row'))
        warehouse_id = data.get('warehouse_id')
        
        if warehouse_id:
            warehouse = get_object_or_404(Warehouse, id=warehouse_id)
        else:
            warehouse = Warehouse.objects.first()
            
        if not warehouse:
             return JsonResponse({'success': False, 'error': 'لم يتم العثور على مستودع'})

        with transaction.atomic():
            # 1. جلب جميع المواقع في هذا الصف مرتبة حسب العمود
            locations = list(Location.objects.filter(
                warehouse=warehouse, 
                row=row_number
            ).order_by('column').prefetch_related('products'))
            
            if not locations:
                return JsonResponse({'success': False, 'error': 'الصف غير موجود'})
            
            # حفظ الحالة السابقة للتراجع
            undo_data = []
            for loc in locations:
                for product in loc.products.all():
                    undo_data.append({
                        'product_id': product.id,
                        'location_id': loc.id
                    })
            
            # حفظ في الجلسة
            request.session['last_compaction_undo'] = {
                'type': 'row',
                'id': row_number,
                'data': undo_data,
                'timestamp': timezone.now().isoformat()
            }

            # 2. تحديد المواقع التي تحتوي على منتجات (المشغولة)
            # نحتاج إلى قائمة بمحتويات المواقع المشغولة
            occupied_contents = []
            
            for loc in locations:
                # نستخدم all() لجلب جميع المنتجات في الموقع
                products = list(loc.products.all())
                if products:
                    occupied_contents.append(products)
            
            # 3. إعادة توزيع المحتويات على المواقع الأولى
            updates_count = 0
            
            # نقل المحتويات إلى المواقع الجديدة
            for i, content_products in enumerate(occupied_contents):
                if i < len(locations):
                    target_loc = locations[i]
                    
                    # التحقق وتحديث الموقع لكل منتج في المجموعة
                    # نفترض أن المنتجات في نفس المجموعة تبقى معاً
                    for product in content_products:
                        if product.location_id != target_loc.id:
                            product.location = target_loc
                            product.save()
                            updates_count += 1
            
            return JsonResponse({
                'success': True, 
                'message': f'تم إعادة ترتيب الصف {row_number} بنجاح.',
                'can_undo': True
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def compact_column(request):
    """ضغط العمود (إزالة الفراغات)"""
    try:
        data = json.loads(request.body)
        col_num = int(data.get('column'))
        
        warehouse = Warehouse.objects.first()
        if not warehouse:
            return JsonResponse({'success': False, 'error': 'المستودع غير موجود'})
        
        with transaction.atomic():
            # الحصول على جميع مواقع العمود مرتبة حسب الصف
            locations = list(Location.objects.filter(
                warehouse=warehouse, 
                column=col_num
            ).order_by('row').prefetch_related('products'))
            
            # حفظ الحالة السابقة للتراجع
            undo_data = []
            for loc in locations:
                for product in loc.products.all():
                    undo_data.append({
                        'product_id': product.id,
                        'location_id': loc.id
                    })
            
            # حفظ في الجلسة
            request.session['last_compaction_undo'] = {
                'type': 'column',
                'id': col_num,
                'data': undo_data,
                'timestamp': timezone.now().isoformat()
            }
            
            # تجميع المنتجات من المواقع المشغولة
            occupied_groups = []
            for loc in locations:
                products = list(loc.products.all())
                if products:
                    occupied_groups.append(products)
            
            # إعادة توزيع المنتجات على المواقع الأولى (من الأعلى للأسفل)
            for i, group in enumerate(occupied_groups):
                target_location = locations[i]
                
                for product in group:
                    if product.location_id != target_location.id:
                        product.location = target_location
                        product.save(update_fields=['location'])
            
        return JsonResponse({'success': True, 'message': f'تم إعادة ترتيب العمود {col_num} بنجاح', 'can_undo': True})
    
    except ValueError:
        return JsonResponse({'success': False, 'error': 'بيانات غير صالحة'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def revert_compaction(request):
    """التراجع عن آخر عملية ترتيب"""
    try:
        undo_info = request.session.get('last_compaction_undo')
        if not undo_info:
            return JsonResponse({'success': False, 'error': 'لا توجد عملية للتراجع عنها'})
        
        undo_data = undo_info.get('data', [])
        if not undo_data:
            return JsonResponse({'success': False, 'error': 'بيانات التراجع فارغة'})
        
        with transaction.atomic():
            # أولاً: نحصل على جميع معرفات المنتجات المتأثرة
            product_ids = [item['product_id'] for item in undo_data]
            
            # نحصل على المنتجات الحالية
            products_map = {p.id: p for p in Product.objects.filter(id__in=product_ids)}
            
            # إعادة المنتجات لمواقعها الأصلية
            for item in undo_data:
                product = products_map.get(item['product_id'])
                if product:
                    product.location_id = item['location_id']
                    product.save(update_fields=['location'])
            
            # مسح بيانات التراجع
            del request.session['last_compaction_undo']
            request.session.modified = True
        
        type_str = "الصف" if undo_info['type'] == 'row' else "العمود"
        return JsonResponse({'success': True, 'message': f'تم التراجع عن ترتيب {type_str} {undo_info["id"]} بنجاح'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})